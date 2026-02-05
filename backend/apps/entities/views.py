"""
API views for entities app.
"""
from datetime import datetime
from io import BytesIO
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse
from django.db.models import Count, Min, Prefetch
from django_filters.rest_framework import DjangoFilterBackend
from .models import (
    ProcessingReport, IFCEntity, PropertySet, SpatialHierarchy,
    IFCType, Material, TypeMapping, TypeDefinitionLayer, MaterialMapping, NS3451Code,
    SemanticType, SemanticTypeIFCMapping,
    TypeBankEntry, TypeBankObservation, TypeBankAlias,
    MaterialLibrary, ProductLibrary, ProductComposition
)
from .serializers import (
    ProcessingReportSerializer, IFCEntitySerializer,
    NS3451CodeSerializer, SemanticTypeSerializer, SemanticTypeListSerializer,
    TypeMappingSerializer, TypeDefinitionLayerSerializer,
    MaterialMappingSerializer, IFCTypeWithMappingSerializer, MaterialWithMappingSerializer,
    TypeBankEntrySerializer, TypeBankEntryListSerializer, TypeBankEntryUpdateSerializer,
    TypeBankObservationSerializer, TypeBankAliasSerializer,
    MaterialLibrarySerializer, MaterialLibraryListSerializer,
    ProductLibrarySerializer, ProductLibraryListSerializer, ProductCompositionSerializer
)
from .services.semantic_normalizer import get_normalizer
from .services.excel_export import export_types_to_excel
from .services.excel_import import import_types_from_excel
from .services.reduzer_export import export_types_to_reduzer


class ProcessingReportViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for viewing processing reports.

    list: Get all processing reports
    retrieve: Get a single processing report

    Filtering:
    - ?model={model_id} - Filter by model
    - ?overall_status=success|partial|failed - Filter by status
    - ?catastrophic_failure=true|false - Filter by catastrophic failures only

    Ordering:
    - ?ordering=-started_at (default: newest first)
    - ?ordering=duration_seconds (sort by duration)
    - ?ordering=total_entities_failed (sort by failure count)
    """
    queryset = ProcessingReport.objects.all()
    serializer_class = ProcessingReportSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['model', 'overall_status', 'catastrophic_failure']
    ordering_fields = ['started_at', 'duration_seconds', 'total_entities_failed']
    ordering = ['-started_at']  # Default: newest first


def get_entity_location(entity):
    """
    Get the full spatial location for an entity.

    Returns dict with:
    - storey_name: Name of the building storey
    - building_name: Name of the building
    - site_name: Name of the site
    - spaces: List of space names the element is in
    """
    location = {
        'storey_name': None,
        'building_name': None,
        'site_name': None,
        'spaces': [],
    }

    if not entity.storey_id:
        return location

    try:
        # Get the storey entity
        storey = IFCEntity.objects.filter(id=entity.storey_id).first()
        if storey:
            location['storey_name'] = storey.name

            # Try to get building and site from SpatialHierarchy
            hierarchy = SpatialHierarchy.objects.filter(
                entity=storey,
                model=entity.model
            ).first()

            if hierarchy and hierarchy.path:
                # Path is array of GUIDs from project to this element
                # Look up each GUID to get names
                path_guids = hierarchy.path

                # Get all entities in the path
                path_entities = IFCEntity.objects.filter(
                    model=entity.model,
                    ifc_guid__in=path_guids
                ).values('ifc_guid', 'ifc_type', 'name')

                # Map GUID to entity data
                guid_to_entity = {e['ifc_guid']: e for e in path_entities}

                for guid in path_guids:
                    if guid in guid_to_entity:
                        ent = guid_to_entity[guid]
                        if ent['ifc_type'] == 'IfcBuilding' or 'Building' in (ent['ifc_type'] or ''):
                            location['building_name'] = ent['name']
                        elif ent['ifc_type'] == 'IfcSite' or 'Site' in (ent['ifc_type'] or ''):
                            location['site_name'] = ent['name']

        # Check for containing spaces (IfcSpace)
        # Look for elements that might contain this entity via IfcRelContainedInSpatialStructure
        # For now, we'll check if there are any IfcSpace entities that reference this element
        # This is done via graph edges if available
        from .models import GraphEdge
        space_edges = GraphEdge.objects.filter(
            model=entity.model,
            target_entity=entity,
            relationship_type='IfcRelContainedInSpatialStructure',
            source_entity__ifc_type='IfcSpace'
        ).select_related('source_entity')

        for edge in space_edges:
            if edge.source_entity.name:
                location['spaces'].append(edge.source_entity.name)

    except Exception as e:
        # Log but don't fail - location is nice-to-have
        print(f"Error getting location for entity {entity.id}: {e}")

    return location


class IFCEntityViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for IFC entities.

    DEPRECATED: This ViewSet is deprecated as of the types-only architecture migration.
    Entity data is no longer stored in the database. The viewer fetches properties
    directly from FastAPI which queries the IFC file.

    This endpoint may return empty results or stale data.
    Use FastAPI /ifc/{file_id}/elements/by-express-id/{express_id} instead.

    list: Get entities (use ?model={id} to filter by model, ?express_id={id} for specific entity)
    retrieve: Get a single entity with full property sets

    Filtering:
    - ?model={model_id} - Filter by model (required for list)
    - ?express_id={express_id} - Filter by express ID (for viewer selection)
    - ?ifc_type={type} - Filter by IFC type (IfcWall, IfcDoor, etc.)
    - ?ifc_guid={guid} - Filter by IFC GUID
    """
    queryset = IFCEntity.objects.all()
    serializer_class = IFCEntitySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['model', 'ifc_type', 'ifc_guid', 'express_id']
    ordering_fields = ['ifc_type', 'name']
    ordering = ['ifc_type', 'name']

    def list(self, request, *args, **kwargs):
        """
        List entities with optional filters.
        For selection by express_id, use the get_by_express_id action instead.
        """
        return super().list(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='by-express-id')
    def get_by_express_id(self, request):
        """
        Get an entity by model ID and express ID (for viewer selection).

        GET /api/entities/by-express-id/?model={model_id}&express_id={express_id}

        Note: Express IDs are NOT stored in the database. This endpoint fetches
        all entities for the model and finds the one at the corresponding index.
        """
        model_id = request.query_params.get('model')
        express_id_str = request.query_params.get('express_id')

        if not model_id or not express_id_str:
            return Response({
                'error': 'Both model and express_id are required'
            }, status=400)

        try:
            express_id = int(express_id_str)
        except ValueError:
            return Response({
                'error': 'express_id must be an integer'
            }, status=400)

        # Look up entity by express_id (stored during IFC parsing)
        entity = None
        fallback_used = False
        try:
            entity = IFCEntity.objects.get(model_id=model_id, express_id=express_id)
        except IFCEntity.DoesNotExist:
            # Fallback for models parsed before express_id was stored
            # Just return first entity - accurate selection requires re-parsing
            entity = IFCEntity.objects.filter(model_id=model_id).first()
            fallback_used = True

        if not entity:
            return Response({
                'error': f'No entities found in model'
            }, status=404)

        # Get property sets
        properties = PropertySet.objects.filter(entity=entity)

        # Group properties by Pset name
        psets = {}
        for prop in properties:
            pset_name = prop.pset_name
            if pset_name not in psets:
                psets[pset_name] = []
            psets[pset_name].append({
                'name': prop.property_name,
                'value': prop.property_value,
                'type': prop.property_type,
            })

        # Get full location info
        location = get_entity_location(entity)

        # Build response
        data = {
            'id': str(entity.id),
            'express_id': entity.express_id,
            'ifc_guid': entity.ifc_guid,
            'ifc_type': entity.ifc_type,
            'predefined_type': entity.predefined_type,
            'object_type': entity.object_type,
            'name': entity.name,
            'description': entity.description,
            'model_id': str(entity.model_id),
            'storey_id': str(entity.storey_id) if entity.storey_id else None,
            # Location (resolved names)
            'storey_name': location['storey_name'],
            'building_name': location['building_name'],
            'site_name': location['site_name'],
            'spaces': location['spaces'],
            # Quantities
            'area': entity.area,
            'volume': entity.volume,
            'length': entity.length,
            'height': entity.height,
            'perimeter': entity.perimeter,
            # Property sets (grouped)
            'property_sets': psets,
        }

        # Add warning if fallback was used (model needs re-parsing)
        if fallback_used:
            data['_warning'] = 'Model missing express_id data. Re-parse for accurate element selection.'

        return Response(data)

    def retrieve(self, request, *args, **kwargs):
        """
        Get a single entity with full property sets grouped by Pset name.
        """
        entity = self.get_object()

        # Get all property sets for this entity
        properties = PropertySet.objects.filter(entity=entity)

        # Group properties by Pset name
        psets = {}
        for prop in properties:
            pset_name = prop.pset_name
            if pset_name not in psets:
                psets[pset_name] = []
            psets[pset_name].append({
                'name': prop.property_name,
                'value': prop.property_value,
                'type': prop.property_type,
            })

        # Get full location info
        location = get_entity_location(entity)

        # Build response
        data = {
            'id': str(entity.id),
            'express_id': entity.express_id,
            'ifc_guid': entity.ifc_guid,
            'ifc_type': entity.ifc_type,
            'predefined_type': entity.predefined_type,
            'object_type': entity.object_type,
            'name': entity.name,
            'description': entity.description,
            'model_id': str(entity.model_id),
            'storey_id': str(entity.storey_id) if entity.storey_id else None,
            # Location (resolved names)
            'storey_name': location['storey_name'],
            'building_name': location['building_name'],
            'site_name': location['site_name'],
            'spaces': location['spaces'],
            # Quantities
            'area': entity.area,
            'volume': entity.volume,
            'length': entity.length,
            'height': entity.height,
            'perimeter': entity.perimeter,
            # Property sets (grouped)
            'property_sets': psets,
        }

        return Response(data)


# =============================================================================
# WAREHOUSE VIEWSETS - NS3451, Types, Materials, Mappings
# =============================================================================

class NS3451CodeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for NS-3451 classification codes (reference data).

    GET /api/ns3451-codes/ - List all codes
    GET /api/ns3451-codes/{code}/ - Get single code
    GET /api/ns3451-codes/?level=2 - Filter by hierarchy level
    GET /api/ns3451-codes/?parent_code=22 - Get children of a code
    GET /api/ns3451-codes/?search=betong - Search by name
    GET /api/ns3451-codes/hierarchy/ - Get nested hierarchy tree
    """
    queryset = NS3451Code.objects.all()
    serializer_class = NS3451CodeSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['level', 'parent_code']
    search_fields = ['code', 'name', 'name_en']
    ordering = ['code']

    @action(detail=False, methods=['get'], url_path='hierarchy')
    def hierarchy(self, request):
        """
        Get NS3451 codes as a nested hierarchy tree for cascading selectors.

        GET /api/ns3451-codes/hierarchy/

        Returns nested structure:
        {
            "2": {
                "code": "2",
                "name": "Bygning",
                "children": {
                    "21": {
                        "code": "21",
                        "name": "Grunn og fundamenter",
                        "children": {...}
                    }
                }
            }
        }
        """
        codes = NS3451Code.objects.all().order_by('code')

        # Build nested hierarchy
        hierarchy = {}

        for code_obj in codes:
            code = code_obj.code
            level = code_obj.level

            node = {
                'code': code,
                'name': code_obj.name,
                'name_en': code_obj.name_en,
                'guidance': code_obj.guidance,
                'level': level,
                'children': {}
            }

            if level == 1:
                # Level 1: Top level
                hierarchy[code] = node
            elif level == 2:
                # Level 2: Parent is first digit
                parent_code = code[0]
                if parent_code in hierarchy:
                    hierarchy[parent_code]['children'][code] = node
            elif level == 3:
                # Level 3: Parent is first 2 digits
                parent_l1 = code[0]
                parent_l2 = code[:2]
                if parent_l1 in hierarchy:
                    if parent_l2 in hierarchy[parent_l1]['children']:
                        hierarchy[parent_l1]['children'][parent_l2]['children'][code] = node
            elif level == 4:
                # Level 4: Parent is first 3 digits
                parent_l1 = code[0]
                parent_l2 = code[:2]
                parent_l3 = code[:3]
                if parent_l1 in hierarchy:
                    if parent_l2 in hierarchy[parent_l1]['children']:
                        if parent_l3 in hierarchy[parent_l1]['children'][parent_l2]['children']:
                            hierarchy[parent_l1]['children'][parent_l2]['children'][parent_l3]['children'][code] = node

        return Response(hierarchy)


class SemanticTypeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for semantic types (PA0802/IFC normalization reference data).

    GET /api/semantic-types/ - List all active semantic types
    GET /api/semantic-types/{id}/ - Get single semantic type with IFC mappings
    GET /api/semantic-types/?category=A-Structural - Filter by category
    GET /api/semantic-types/?search=beam - Search by name

    Actions:
    GET /api/semantic-types/by-category/ - Get types grouped by category
    GET /api/semantic-types/for-ifc-class/?ifc_class=IfcBeamType - Get types for IFC class
    """
    queryset = SemanticType.objects.filter(is_active=True).prefetch_related('ifc_mappings')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'is_active', 'canonical_ifc_class']
    search_fields = ['code', 'name_no', 'name_en', 'description']
    ordering = ['category', 'code']

    def get_serializer_class(self):
        if self.action == 'list':
            return SemanticTypeListSerializer
        return SemanticTypeSerializer

    @action(detail=False, methods=['get'], url_path='by-category')
    def by_category(self, request):
        """
        Get semantic types grouped by category.

        GET /api/semantic-types/by-category/

        Returns:
        {
            "A-Structural": [
                {"code": "AB", "name_en": "Beam", ...},
                {"code": "AS", "name_en": "Column", ...}
            ],
            "D-Openings": [...]
        }
        """
        types = self.get_queryset()
        grouped = {}

        for st in types:
            if st.category not in grouped:
                grouped[st.category] = []
            grouped[st.category].append(SemanticTypeListSerializer(st).data)

        return Response(grouped)

    @action(detail=False, methods=['get'], url_path='for-ifc-class')
    def for_ifc_class(self, request):
        """
        Get semantic types that match a given IFC class.

        GET /api/semantic-types/for-ifc-class/?ifc_class=IfcBeamType

        Returns list of matching semantic types with mapping info.
        """
        ifc_class = request.query_params.get('ifc_class')
        if not ifc_class:
            return Response(
                {'error': 'ifc_class query parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Normalize IFC class (remove Type suffix)
        normalized_class = ifc_class
        if ifc_class.endswith('Type'):
            normalized_class = ifc_class[:-4]

        # Find mappings for this IFC class
        mappings = SemanticTypeIFCMapping.objects.filter(
            ifc_class__in=[ifc_class, normalized_class]
        ).select_related('semantic_type').order_by('-is_primary', '-confidence_hint')

        result = []
        seen_codes = set()

        for mapping in mappings:
            if mapping.semantic_type.code not in seen_codes:
                result.append({
                    'code': mapping.semantic_type.code,
                    'name_no': mapping.semantic_type.name_no,
                    'name_en': mapping.semantic_type.name_en,
                    'category': mapping.semantic_type.category,
                    'is_primary': mapping.is_primary,
                    'is_common_misuse': mapping.is_common_misuse,
                    'confidence_hint': mapping.confidence_hint,
                    'note': mapping.note,
                })
                seen_codes.add(mapping.semantic_type.code)

        return Response(result)


class IFCTypeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for IFC types with mapping status.

    GET /api/types/?model={id} - List types for a model (required)
    GET /api/types/{id}/ - Get single type with mapping
    GET /api/types/?model={id}&ifc_type=IfcWallType - Filter by IFC class

    Returns instance_count (how many entities use this type) and
    mapping object (NS3451 code, status, etc.)

    Note: Pagination disabled - types are small and mapping workflow needs all of them.
    Typical models have 50-500 types; even 1000+ types is fine to return at once.
    """
    queryset = IFCType.objects.select_related(
        'mapping', 'mapping__ns3451'
    ).prefetch_related(
        Prefetch('mapping__definition_layers', queryset=TypeDefinitionLayer.objects.order_by('layer_order'))
    ).all()
    serializer_class = IFCTypeWithMappingSerializer
    pagination_class = None  # Return all types - needed for mapping workflow
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['model', 'ifc_type']
    search_fields = ['type_name', 'ifc_type']
    ordering = ['ifc_type', 'type_name']

    def get_queryset(self):
        """
        Filter types by default to exclude those with 0 instances.

        Use ?include_unused=true to show all types including unused ones.
        This reduces noise from template types that have no instances in the model.
        """
        qs = super().get_queryset()

        include_unused = self.request.query_params.get('include_unused', 'false').lower() == 'true'
        if not include_unused:
            # Use stored instance_count field (populated during parsing)
            qs = qs.filter(instance_count__gt=0)

        return qs

    @action(detail=False, methods=['get'], url_path='summary')
    def mapping_summary(self, request):
        """
        Get mapping summary for a model.

        GET /api/types/summary/?model={id}

        Returns counts: total, mapped, pending, ignored, review
        """
        model_id = request.query_params.get('model')
        if not model_id:
            return Response({'error': 'model parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        types = IFCType.objects.filter(model_id=model_id)
        total = types.count()

        # Count by mapping status
        mapped = types.filter(mapping__mapping_status='mapped').count()
        pending = types.filter(mapping__mapping_status='pending').count()
        ignored = types.filter(mapping__mapping_status='ignored').count()
        review = types.filter(mapping__mapping_status='review').count()
        unmapped = total - (mapped + pending + ignored + review)

        return Response({
            'total': total,
            'mapped': mapped,
            'pending': pending + unmapped,  # Treat unmapped as pending
            'ignored': ignored,
            'review': review,
            'progress_percent': round((mapped / total * 100) if total > 0 else 0, 1)
        })

    @action(detail=False, methods=['get'], url_path='consolidated')
    def consolidated(self, request):
        """
        Get types consolidated by signature (ifc_type + type_name + key properties).

        GET /api/entities/types/consolidated/?model={id}

        Creates a "type signature" for ML-style grouping:
        - ifc_type: IfcColumnType
        - type_name: CFRS 150x150x10
        - is_external: False
        - load_bearing: True
        - material: S355

        Two types with same name but different key properties = different consolidated types.
        """
        model_id = request.query_params.get('model')
        if not model_id:
            return Response({'error': 'model parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        from .models import TypeAssignment
        from collections import defaultdict

        # Key properties for type signature (order matters for consistency)
        KEY_PROPERTIES = ['IsExternal', 'LoadBearing', 'FireRating', 'Reference']

        # First pass: group by (ifc_type, type_name) to get candidate groups
        type_groups = IFCType.objects.filter(model_id=model_id).values(
            'ifc_type', 'type_name'
        ).annotate(guid_count=Count('id')).order_by('ifc_type', 'type_name')

        results = []
        for group in type_groups:
            # Get all types in this name group
            types_in_group = IFCType.objects.filter(
                model_id=model_id,
                ifc_type=group['ifc_type'],
                type_name=group['type_name']
            )
            type_ids = list(types_in_group.values_list('id', flat=True))

            # Get entities for these types
            entity_ids = list(TypeAssignment.objects.filter(
                type_id__in=type_ids
            ).values_list('entity_id', flat=True))

            # Extract key property values (most common value for each property)
            signature_props = {}
            for prop_name in KEY_PROPERTIES:
                values = list(PropertySet.objects.filter(
                    entity_id__in=entity_ids,
                    property_name=prop_name
                ).values_list('property_value', flat=True))
                if values:
                    # Use most common value as the signature value
                    from collections import Counter
                    most_common = Counter(values).most_common(1)
                    signature_props[prop_name.lower()] = most_common[0][0] if most_common else None

            # Get representative type for mapping info
            rep_type = types_in_group.select_related('mapping', 'mapping__ns3451').first()

            # Get material from entities (common property)
            materials = list(PropertySet.objects.filter(
                entity_id__in=entity_ids,
                property_name__in=['Structural Material', 'Material']
            ).values_list('property_value', flat=True).distinct()[:3])

            # Get mapping info
            mapping_status = None
            ns3451_code = None
            ns3451_name = None
            if rep_type and hasattr(rep_type, 'mapping') and rep_type.mapping:
                mapping_status = rep_type.mapping.mapping_status
                ns3451_code = rep_type.mapping.ns3451_code
                if rep_type.mapping.ns3451:
                    ns3451_name = rep_type.mapping.ns3451.name

            results.append({
                'ifc_type': group['ifc_type'],
                'type_name': group['type_name'],
                'guid_count': group['guid_count'],
                'instance_count': len(entity_ids),
                'representative_id': str(rep_type.id) if rep_type else None,
                # Key properties for ML signature
                'is_external': signature_props.get('isexternal'),
                'load_bearing': signature_props.get('loadbearing'),
                'fire_rating': signature_props.get('firerating'),
                'reference': signature_props.get('reference'),
                'materials': materials,
                # Mapping info
                'mapping_status': mapping_status,
                'ns3451_code': ns3451_code,
                'ns3451_name': ns3451_name,
            })

        return Response({
            'model_id': model_id,
            'consolidated_count': len(results),
            'raw_type_count': IFCType.objects.filter(model_id=model_id).count(),
            'types': results
        })

    @action(detail=False, methods=['post'], url_path='map-consolidated')
    def map_consolidated(self, request):
        """
        Apply mapping to ALL types matching (ifc_type, type_name) at once.

        POST /api/entities/types/map-consolidated/
        {
            "model_id": "uuid",
            "ifc_type": "IfcColumnType",
            "type_name": "CFRS 150x150x10",
            "ns3451_code": "234",
            "mapping_status": "mapped",
            "representative_unit": "stk",
            "notes": "Steel columns"
        }

        This finds all IFCType records matching (model_id, ifc_type, type_name)
        and creates/updates TypeMapping for each one.

        Returns:
        {
            "success": true,
            "types_updated": 37,
            "mapping_data": { ... }
        }
        """
        model_id = request.data.get('model_id')
        ifc_type = request.data.get('ifc_type')
        type_name = request.data.get('type_name')

        if not all([model_id, ifc_type, type_name]):
            return Response(
                {'error': 'model_id, ifc_type, and type_name are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Find all matching types
        matching_types = IFCType.objects.filter(
            model_id=model_id,
            ifc_type=ifc_type,
            type_name=type_name
        )

        if not matching_types.exists():
            return Response(
                {'error': f'No types found matching {ifc_type} / {type_name}'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Extract mapping fields from request
        mapping_fields = {}
        for field in ['ns3451_code', 'mapping_status', 'representative_unit', 'notes', 'mapped_by']:
            if field in request.data:
                mapping_fields[field] = request.data[field]

        # Default status to 'mapped' if ns3451_code provided
        if 'ns3451_code' in mapping_fields and mapping_fields['ns3451_code']:
            mapping_fields.setdefault('mapping_status', 'mapped')
            mapping_fields['mapped_at'] = datetime.now()

        # Create/update TypeMapping for each matching type
        updated_count = 0
        created_count = 0

        for ifc_type_obj in matching_types:
            mapping, created = TypeMapping.objects.update_or_create(
                ifc_type=ifc_type_obj,
                defaults=mapping_fields
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        return Response({
            'success': True,
            'types_matched': matching_types.count(),
            'mappings_created': created_count,
            'mappings_updated': updated_count,
            'mapping_data': mapping_fields
        })

    @action(detail=True, methods=['get'], url_path='instances')
    def instances(self, request, pk=None):
        """
        Get entity instances for a type (for viewer navigation).

        GET /api/types/{id}/instances/

        Queries the IFC file directly via FastAPI to get instance GUIDs.
        Used by the Instance Viewer to navigate through type instances.
        """
        import httpx
        from django.conf import settings

        ifc_type = self.get_object()

        # Get the model's IFC file URL
        model = ifc_type.model
        if not model.file_url:
            return Response({
                'type_id': str(ifc_type.id),
                'type_name': ifc_type.type_name,
                'type_guid': ifc_type.type_guid,
                'ifc_type': ifc_type.ifc_type,
                'model_id': str(model.id),
                'total_count': 0,
                'instances': [],
                'error': 'No IFC file available for this model'
            })

        # Get FastAPI base URL from settings (matches IFCServiceClient)
        fastapi_url = getattr(settings, 'IFC_SERVICE_URL', 'http://localhost:8001')

        try:
            with httpx.Client(timeout=60.0) as client:
                # Step 1: Load the IFC file into FastAPI (idempotent - returns cached file_id)
                open_response = client.post(
                    f"{fastapi_url}/api/v1/ifc/open/url",
                    json={"file_url": model.file_url}
                )
                open_response.raise_for_status()
                file_id = open_response.json()['file_id']

                # Step 2: Get type instances
                instances_response = client.get(
                    f"{fastapi_url}/api/v1/ifc/{file_id}/types/{ifc_type.type_guid}/instances"
                )
                instances_response.raise_for_status()
                data = instances_response.json()

                return Response({
                    'type_id': str(ifc_type.id),
                    'type_name': ifc_type.type_name,
                    'type_guid': ifc_type.type_guid,
                    'ifc_type': ifc_type.ifc_type,
                    'model_id': str(model.id),
                    'total_count': data['total_count'],
                    'instances': data['instances']
                })

        except httpx.HTTPStatusError as e:
            return Response({
                'type_id': str(ifc_type.id),
                'type_name': ifc_type.type_name,
                'type_guid': ifc_type.type_guid,
                'ifc_type': ifc_type.ifc_type,
                'model_id': str(model.id),
                'total_count': 0,
                'instances': [],
                'error': f'FastAPI error: {e.response.status_code} - {e.response.text}'
            }, status=e.response.status_code)
        except Exception as e:
            return Response({
                'type_id': str(ifc_type.id),
                'type_name': ifc_type.type_name,
                'type_guid': ifc_type.type_guid,
                'ifc_type': ifc_type.ifc_type,
                'model_id': str(model.id),
                'total_count': 0,
                'instances': [],
                'error': f'Failed to query IFC file: {str(e)}'
            }, status=500)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Export types to Excel template for batch mapping.

        GET /api/types/export-excel/?model={id}

        Returns Excel file with:
        - Editable columns (A-D): NS3451 Code, Unit, Notes, Status
        - Read-only columns (E-R): Type metadata and aggregated properties
        - Rows grouped by IfcEntity
        """
        model_id = request.query_params.get('model')
        if not model_id:
            return Response(
                {'error': 'model parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check model exists and has types
        type_count = IFCType.objects.filter(model_id=model_id).count()
        if type_count == 0:
            return Response(
                {'error': 'No types found for this model'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            # Get model name for filename
            from apps.models.models import Model
            model = Model.objects.filter(id=model_id).first()
            model_name = model.name if model else 'unknown'

            # Generate Excel file
            excel_buffer = export_types_to_excel(model_id)

            # Build filename
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f"types_{model_name}_{date_str}.xlsx"
            # Sanitize filename
            filename = "".join(c for c in filename if c.isalnum() or c in '._-')

            # Return as file download
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response(
                {'error': f'Failed to generate Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='export-reduzer')
    def export_reduzer(self, request):
        """
        Export types to Reduzer-compatible Excel format for LCA import.

        GET /api/entities/types/export-reduzer/?model={id}&include_unmapped=false

        Returns Excel file with Reduzer import format:
        - description: Type name
        - NS3451:2009: Building classification code
        - quantity: Aggregated quantity from instances
        - unit: Unit of measurement (stk, m, m2, m3)
        - component: Discipline or NS3451 category grouping
        - productIDType: Product ID type (EPD, NOBB, etc.)
        - productID: Product/EPD identifier
        - notes: Additional notes

        Only includes types with NS3451 mapping unless include_unmapped=true.
        Types with material layers expand to multiple rows (one per EPD).
        """
        model_id = request.query_params.get('model')
        include_unmapped = request.query_params.get('include_unmapped', 'false').lower() == 'true'

        if not model_id:
            return Response(
                {'error': 'model parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check model exists
        from apps.models.models import Model
        model = Model.objects.filter(id=model_id).first()
        if not model:
            return Response(
                {'error': 'Model not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check for mapped types
        if not include_unmapped:
            mapped_count = IFCType.objects.filter(
                model_id=model_id,
                mapping__ns3451_code__isnull=False
            ).exclude(mapping__ns3451_code='').count()

            if mapped_count == 0:
                return Response(
                    {'error': 'No mapped types found. Map types to NS3451 codes first or use include_unmapped=true'},
                    status=status.HTTP_404_NOT_FOUND
                )

        try:
            # Generate Reduzer Excel file
            excel_buffer = export_types_to_reduzer(model_id, include_unmapped=include_unmapped)

            # Build filename
            date_str = datetime.now().strftime('%Y-%m-%d')
            model_name = model.name if model else 'unknown'
            filename = f"reduzer_{model_name}_{date_str}.xlsx"
            # Sanitize filename
            filename = "".join(c for c in filename if c.isalnum() or c in '._-')

            # Return as file download
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response(
                {'error': f'Failed to generate Reduzer export: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='import-excel', parser_classes=[MultiPartParser])
    def import_excel(self, request):
        """
        Import type mappings from Excel file.

        POST /api/types/import-excel/
        Body (multipart/form-data):
        - file: Excel file (.xlsx)
        - model_id: UUID of the model

        Returns:
        {
            "success": true,
            "summary": {"total_rows": 45, "updated": 42, "created": 0, "skipped": 2, "error_count": 1},
            "errors": [{"row": 15, "type_guid": "abc123", "error": "Invalid NS3451 code: 999"}],
            "warnings": [{"row": 8, "type_guid": "def456", "warning": "NS3451 code empty, status set to pending"}]
        }
        """
        file = request.FILES.get('file')
        model_id = request.data.get('model_id')

        if not file:
            return Response(
                {'error': 'file is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not model_id:
            return Response(
                {'error': 'model_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate file type
        if not file.name.endswith('.xlsx'):
            return Response(
                {'error': 'File must be an Excel file (.xlsx)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check model exists
        type_count = IFCType.objects.filter(model_id=model_id).count()
        if type_count == 0:
            return Response(
                {'error': 'No types found for this model'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            # Read file content
            file_content = BytesIO(file.read())

            # Get username if authenticated
            username = None
            if request.user and request.user.is_authenticated:
                username = request.user.username or request.user.email

            # Import
            result = import_types_from_excel(model_id, file_content, username)

            return Response(result.to_dict())

        except Exception as e:
            return Response(
                {'error': f'Failed to import Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='dashboard-metrics')
    def dashboard_metrics(self, request):
        """
        Get dashboard metrics for a project or specific model.

        GET /api/types/dashboard-metrics/?project_id={id}
        GET /api/types/dashboard-metrics/?model_id={id}

        Returns:
        - project_summary: Overall health score, status counts
        - models: Per-model breakdown with health scores
        - by_discipline: Aggregated by discipline (ARK, RIB, etc.)

        Health Score Formula (0-100):
        - Classification score (40%): types with NS3451 code
        - Unit score (20%): types with representative_unit
        - Material score (40%): types with at least 1 material layer with quantity

        Status thresholds:
        - healthy (green): >= 80
        - warning (yellow): >= 50
        - critical (red): < 50
        """
        from django.db.models import Count, Q, Exists, OuterRef
        from apps.models.models import Model

        project_id = request.query_params.get('project_id')
        model_id = request.query_params.get('model_id')

        if not project_id and not model_id:
            return Response(
                {'error': 'project_id or model_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        def calculate_health_score(types_qs):
            """Calculate health score for a queryset of types."""
            total = types_qs.count()
            if total == 0:
                return {
                    'total': 0,
                    'health_score': 0,
                    'status': 'critical',
                    'classification_score': 0,
                    'unit_score': 0,
                    'material_score': 0,
                }

            # Classification score (40%): types with NS3451 code
            with_ns3451 = types_qs.filter(
                mapping__ns3451_code__isnull=False
            ).exclude(mapping__ns3451_code='').count()
            classification_score = (with_ns3451 / total) * 100

            # Unit score (20%): types with representative_unit
            with_unit = types_qs.filter(
                mapping__representative_unit__isnull=False
            ).exclude(mapping__representative_unit='').count()
            unit_score = (with_unit / total) * 100

            # Material score (40%): types with at least 1 material layer with quantity > 0
            # Use subquery to check if any definition_layer exists with quantity_per_unit > 0
            has_material_layer = TypeDefinitionLayer.objects.filter(
                type_mapping=OuterRef('mapping'),
                quantity_per_unit__gt=0
            )
            with_materials = types_qs.filter(
                mapping__isnull=False
            ).annotate(
                has_layers=Exists(has_material_layer)
            ).filter(has_layers=True).count()
            material_score = (with_materials / total) * 100

            # Composite health score
            health_score = round(
                classification_score * 0.4 +
                unit_score * 0.2 +
                material_score * 0.4,
                1
            )

            # Status threshold
            if health_score >= 80:
                health_status = 'healthy'
            elif health_score >= 50:
                health_status = 'warning'
            else:
                health_status = 'critical'

            return {
                'total_types': total,
                'health_score': health_score,
                'status': health_status,
                'classification_percent': round(classification_score, 1),
                'unit_percent': round(unit_score, 1),
                'material_percent': round(material_score, 1),
            }

        def get_status_counts(types_qs):
            """Get mapping status counts."""
            total = types_qs.count()
            mapped = types_qs.filter(mapping__mapping_status='mapped').count()
            pending = types_qs.filter(
                Q(mapping__mapping_status='pending') | Q(mapping__isnull=True)
            ).count()
            ignored = types_qs.filter(mapping__mapping_status='ignored').count()
            review = types_qs.filter(mapping__mapping_status='review').count()
            followup = types_qs.filter(mapping__mapping_status='followup').count()

            return {
                'total': total,
                'mapped': mapped,
                'pending': pending,
                'ignored': ignored,
                'review': review,
                'followup': followup,
                'progress_percent': round((mapped / total * 100) if total > 0 else 0, 1),
            }

        # Single model mode
        if model_id:
            types_qs = IFCType.objects.filter(model_id=model_id, instance_count__gt=0)
            model = Model.objects.filter(id=model_id).first()

            if not model:
                return Response({'error': 'Model not found'}, status=status.HTTP_404_NOT_FOUND)

            health = calculate_health_score(types_qs)
            counts = get_status_counts(types_qs)

            return Response({
                'mode': 'model',
                'model_id': model_id,
                'model_name': model.name,
                **health,
                **counts,
            })

        # Project mode - aggregate across all models
        models = Model.objects.filter(project_id=project_id)
        if not models.exists():
            return Response({'error': 'No models found for this project'}, status=status.HTTP_404_NOT_FOUND)

        # Project-level aggregation
        all_types = IFCType.objects.filter(
            model__project_id=project_id,
            instance_count__gt=0
        )
        project_health = calculate_health_score(all_types)
        project_counts = get_status_counts(all_types)

        # Per-model breakdown
        model_breakdown = []
        for model in models:
            model_types = IFCType.objects.filter(model=model, instance_count__gt=0)
            if model_types.count() == 0:
                continue  # Skip models with no types

            m_health = calculate_health_score(model_types)
            m_counts = get_status_counts(model_types)

            # Get discipline from model name or mapping
            discipline = None
            if model.name:
                # Try to extract discipline from model name (e.g., "ARK_Model.ifc")
                for disc in ['ARK', 'RIB', 'RIV', 'RIE', 'RIBE', 'RIRV']:
                    if disc in model.name.upper():
                        discipline = disc
                        break

            model_breakdown.append({
                'id': str(model.id),
                'name': model.name,
                'discipline': discipline,
                'total_types': m_health['total_types'],
                'mapped': m_counts['mapped'],
                'pending': m_counts['pending'],
                'ignored': m_counts['ignored'],
                'review': m_counts['review'],
                'followup': m_counts['followup'],
                'health_score': m_health['health_score'],
                'status': m_health['status'],
            })

        # Sort by health score (worst first for attention)
        model_breakdown.sort(key=lambda x: x['health_score'])

        # By discipline aggregation
        by_discipline = {}
        for model_data in model_breakdown:
            disc = model_data['discipline'] or 'Unknown'
            if disc not in by_discipline:
                by_discipline[disc] = {
                    'total': 0,
                    'mapped': 0,
                    'model_count': 0,
                    'health_scores': [],
                }
            by_discipline[disc]['total'] += model_data['total_types']
            by_discipline[disc]['mapped'] += model_data['mapped']
            by_discipline[disc]['model_count'] += 1
            by_discipline[disc]['health_scores'].append(model_data['health_score'])

        # Calculate average health score per discipline
        for disc, data in by_discipline.items():
            if data['health_scores']:
                data['health_score'] = round(
                    sum(data['health_scores']) / len(data['health_scores']), 1
                )
            else:
                data['health_score'] = 0
            del data['health_scores']

        return Response({
            'mode': 'project',
            'project_id': project_id,
            'project_summary': {
                **project_health,
                **project_counts,
            },
            'models': model_breakdown,
            'by_discipline': by_discipline,
        })


class TypeMappingViewSet(viewsets.ModelViewSet):
    """
    API endpoint for type → NS3451 mappings.

    GET /api/type-mappings/?ifc_type__model={id} - List mappings for a model
    POST /api/type-mappings/ - Create mapping
    PATCH /api/type-mappings/{id}/ - Update mapping (ns3451_code, status)
    DELETE /api/type-mappings/{id}/ - Delete mapping

    To map a type:
    1. Find or create TypeMapping for the IFCType
    2. PATCH with ns3451_code and mapping_status='mapped'
    """
    queryset = TypeMapping.objects.select_related('ifc_type', 'ns3451').all()
    serializer_class = TypeMappingSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['ifc_type', 'ifc_type__model', 'mapping_status', 'ns3451_code']
    ordering = ['-updated_at']

    def perform_update(self, serializer):
        """Set mapped_at timestamp when status changes to mapped."""
        instance = serializer.save()
        if instance.mapping_status == 'mapped' and not instance.mapped_at:
            instance.mapped_at = datetime.now()
            instance.save(update_fields=['mapped_at'])

    @action(detail=False, methods=['post'], url_path='bulk-update')
    def bulk_update(self, request):
        """
        Bulk update multiple type mappings.

        POST /api/type-mappings/bulk-update/
        Body: {
            "mappings": [
                {"ifc_type_id": "uuid", "ns3451_code": "222", "mapping_status": "mapped"},
                ...
            ]
        }
        """
        mappings_data = request.data.get('mappings', [])
        if not mappings_data:
            return Response({'error': 'mappings array is required'}, status=status.HTTP_400_BAD_REQUEST)

        updated = 0
        created = 0
        errors = []

        for item in mappings_data:
            ifc_type_id = item.get('ifc_type_id')
            ns3451_code = item.get('ns3451_code')
            mapping_status = item.get('mapping_status', 'mapped')

            try:
                mapping, was_created = TypeMapping.objects.update_or_create(
                    ifc_type_id=ifc_type_id,
                    defaults={
                        'ns3451_code': ns3451_code,
                        'ns3451_id': ns3451_code,  # FK uses code as PK
                        'mapping_status': mapping_status,
                        'mapped_at': datetime.now() if mapping_status == 'mapped' else None,
                    }
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({'ifc_type_id': ifc_type_id, 'error': str(e)})

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors
        })


class TypeDefinitionLayerViewSet(viewsets.ModelViewSet):
    """
    API endpoint for type definition layers (material composition).

    GET /api/type-definition-layers/?type_mapping={id} - List layers for a type mapping
    POST /api/type-definition-layers/ - Create layer
    PATCH /api/type-definition-layers/{id}/ - Update layer
    DELETE /api/type-definition-layers/{id}/ - Delete layer

    POST /api/type-definition-layers/bulk-update/ - Bulk create/update layers for a type

    Layers define the material composition of a type (e.g., wall layers:
    exterior cladding, insulation, structure, interior finish).
    """
    queryset = TypeDefinitionLayer.objects.select_related('type_mapping').all()
    serializer_class = TypeDefinitionLayerSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['type_mapping']
    ordering = ['type_mapping', 'layer_order']

    @action(detail=False, methods=['post'], url_path='bulk-update')
    def bulk_update(self, request):
        """
        Bulk create/update layers for a type mapping.

        POST /api/type-definition-layers/bulk-update/
        Body: {
            "type_mapping_id": "uuid",
            "layers": [
                {"layer_order": 1, "material_name": "Gypsum board", "thickness_mm": 12.5, "epd_id": null},
                {"layer_order": 2, "material_name": "Mineral wool", "thickness_mm": 150, "epd_id": "EPD-123"},
                {"layer_order": 3, "material_name": "Brick", "thickness_mm": 108, "epd_id": null}
            ]
        }

        This will:
        1. Delete existing layers for the type mapping
        2. Create new layers from the provided data
        """
        type_mapping_id = request.data.get('type_mapping_id')
        layers_data = request.data.get('layers', [])

        if not type_mapping_id:
            return Response({'error': 'type_mapping_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            type_mapping = TypeMapping.objects.get(id=type_mapping_id)
        except TypeMapping.DoesNotExist:
            return Response({'error': 'TypeMapping not found'}, status=status.HTTP_404_NOT_FOUND)

        # Delete existing layers
        TypeDefinitionLayer.objects.filter(type_mapping=type_mapping).delete()

        # Create new layers
        created_layers = []
        for layer_data in layers_data:
            layer = TypeDefinitionLayer.objects.create(
                type_mapping=type_mapping,
                layer_order=layer_data.get('layer_order', 1),
                material_name=layer_data.get('material_name', ''),
                thickness_mm=layer_data.get('thickness_mm', 0),
                epd_id=layer_data.get('epd_id'),
                notes=layer_data.get('notes', ''),
            )
            created_layers.append(TypeDefinitionLayerSerializer(layer).data)

        return Response({
            'type_mapping_id': str(type_mapping_id),
            'layers': created_layers,
            'count': len(created_layers)
        })


class MaterialViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for materials with mapping status.

    GET /api/materials/?model={id} - List materials for a model (required)
    GET /api/materials/{id}/ - Get single material with mapping
    """
    queryset = Material.objects.select_related('mapping').all()
    serializer_class = MaterialWithMappingSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['model', 'category']
    search_fields = ['name', 'category']
    ordering = ['name']

    @action(detail=False, methods=['get'], url_path='summary')
    def mapping_summary(self, request):
        """
        Get mapping summary for a model.

        GET /api/materials/summary/?model={id}
        """
        model_id = request.query_params.get('model')
        if not model_id:
            return Response({'error': 'model parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        materials = Material.objects.filter(model_id=model_id)
        total = materials.count()

        mapped = materials.filter(mapping__mapping_status='mapped').count()
        pending = materials.filter(mapping__mapping_status='pending').count()
        ignored = materials.filter(mapping__mapping_status='ignored').count()
        review = materials.filter(mapping__mapping_status='review').count()
        unmapped = total - (mapped + pending + ignored + review)

        return Response({
            'total': total,
            'mapped': mapped,
            'pending': pending + unmapped,
            'ignored': ignored,
            'review': review,
            'progress_percent': round((mapped / total * 100) if total > 0 else 0, 1)
        })


class MaterialMappingViewSet(viewsets.ModelViewSet):
    """
    API endpoint for material mappings.

    GET /api/material-mappings/?material__model={id} - List mappings for a model
    POST /api/material-mappings/ - Create mapping
    PATCH /api/material-mappings/{id}/ - Update mapping
    DELETE /api/material-mappings/{id}/ - Delete mapping
    """
    queryset = MaterialMapping.objects.select_related('material').all()
    serializer_class = MaterialMappingSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['material', 'material__model', 'mapping_status']
    ordering = ['-updated_at']

    def perform_update(self, serializer):
        """Set mapped_at timestamp when status changes to mapped."""
        instance = serializer.save()
        if instance.mapping_status == 'mapped' and not instance.mapped_at:
            instance.mapped_at = datetime.now()
            instance.save(update_fields=['mapped_at'])


# =============================================================================
# TYPE BANK VIEWSETS - Global Type Classification System
# =============================================================================

class TypeBankEntryViewSet(viewsets.ModelViewSet):
    """
    API endpoint for TypeBank entries (global type classification system).

    TypeBank is a shared type library that:
    - Classifies types based on (ifc_class, type_name, predefined_type, material)
    - Tracks where types have been observed across models
    - Stores expert labels (NS3451, discipline, canonical name)
    - REPLACES the per-model TypeMapping system

    List endpoints:
    GET /api/type-bank/ - List all types (paginated, filterable)
    GET /api/type-bank/?ifc_class=IfcWallType - Filter by IFC class
    GET /api/type-bank/?mapping_status=pending - Filter by mapping status
    GET /api/type-bank/?ns3451_code=222 - Filter by NS3451 code
    GET /api/type-bank/?search=concrete - Search by type_name or canonical_name

    Detail endpoints:
    GET /api/type-bank/{id}/ - Get full entry with observations and aliases
    PATCH /api/type-bank/{id}/ - Update classification labels

    Actions:
    GET /api/type-bank/summary/ - Get mapping progress stats
    GET /api/type-bank/export-excel/ - Export to Excel template
    GET /api/type-bank/export-json/ - Export to JSON for ML training
    POST /api/type-bank/{id}/add-alias/ - Add alias to entry
    POST /api/type-bank/{id}/merge-into/ - Merge another entry into this one
    """

    queryset = TypeBankEntry.objects.select_related('ns3451').annotate(
        _observation_count=Count('observations')
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ifc_class', 'mapping_status', 'ns3451_code', 'discipline', 'confidence']
    search_fields = ['type_name', 'canonical_name', 'material', 'ifc_class']
    ordering_fields = ['ifc_class', 'type_name', 'total_instance_count', 'source_model_count', 'updated_at']
    ordering = ['ifc_class', 'type_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return TypeBankEntryListSerializer
        elif self.action in ['update', 'partial_update']:
            return TypeBankEntryUpdateSerializer
        return TypeBankEntrySerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        # For detail views, prefetch observations and aliases
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related(
                Prefetch('observations', queryset=TypeBankObservation.objects.select_related(
                    'source_model', 'source_model__project', 'source_type'
                ).order_by('-observed_at')),
                'aliases'
            )
        return queryset

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        Get TypeBank mapping progress summary.

        GET /api/type-bank/summary/

        Returns:
        {
            "total": 1234,
            "mapped": 890,
            "pending": 300,
            "ignored": 20,
            "review": 24,
            "progress_percent": 72.1,
            "by_ifc_class": {"IfcWallType": 150, "IfcColumnType": 80, ...},
            "by_discipline": {"ARK": 400, "RIB": 300, ...}
        }
        """
        total = TypeBankEntry.objects.count()
        mapped = TypeBankEntry.objects.filter(mapping_status='mapped').count()
        pending = TypeBankEntry.objects.filter(mapping_status='pending').count()
        ignored = TypeBankEntry.objects.filter(mapping_status='ignored').count()
        review = TypeBankEntry.objects.filter(mapping_status='review').count()

        # By IFC class
        by_class = dict(
            TypeBankEntry.objects.values('ifc_class')
            .annotate(count=Count('id'))
            .values_list('ifc_class', 'count')
        )

        # By discipline (only for mapped entries)
        by_discipline = dict(
            TypeBankEntry.objects.filter(discipline__isnull=False)
            .exclude(discipline='')
            .values('discipline')
            .annotate(count=Count('id'))
            .values_list('discipline', 'count')
        )

        return Response({
            'total': total,
            'mapped': mapped,
            'pending': pending,
            'ignored': ignored,
            'review': review,
            'progress_percent': round((mapped / total * 100) if total > 0 else 0, 1),
            'by_ifc_class': by_class,
            'by_discipline': by_discipline,
        })

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Export TypeBank to Excel for batch editing.

        GET /api/type-bank/export-excel/
        GET /api/type-bank/export-excel/?mapping_status=pending - Only pending types
        GET /api/type-bank/export-excel/?ifc_class=IfcWallType - Filter by class

        Returns Excel file with editable columns:
        - NS3451 Code, Discipline, Canonical Name, Representative Unit, Status, Notes
        And read-only columns:
        - IFC Class, Type Name, Predefined Type, Material, Instance Count, Model Count
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # Apply filters
        queryset = self.filter_queryset(self.get_queryset())

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TypeBank"

        # Headers
        headers = [
            # Editable columns (A-F)
            ('NS3451 Code', 12),
            ('Discipline', 10),
            ('Canonical Name', 30),
            ('Unit', 8),
            ('Status', 12),
            ('Notes', 40),
            # Read-only columns (G-N)
            ('ID', 36),
            ('IFC Class', 20),
            ('Type Name', 40),
            ('Predefined Type', 15),
            ('Material', 30),
            ('Instance Count', 14),
            ('Model Count', 12),
            ('Confidence', 12),
        ]

        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        readonly_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        for col, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[cell.column_letter].width = width

        # Data rows
        for row_num, entry in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=entry.ns3451_code or '')
            ws.cell(row=row_num, column=2, value=entry.discipline or '')
            ws.cell(row=row_num, column=3, value=entry.canonical_name or '')
            ws.cell(row=row_num, column=4, value=entry.representative_unit or '')
            ws.cell(row=row_num, column=5, value=entry.mapping_status)
            ws.cell(row=row_num, column=6, value=entry.notes or '')

            # Read-only columns (grey background)
            for col in range(7, 15):
                ws.cell(row=row_num, column=col).fill = readonly_fill

            ws.cell(row=row_num, column=7, value=str(entry.id))
            ws.cell(row=row_num, column=8, value=entry.ifc_class)
            ws.cell(row=row_num, column=9, value=entry.type_name)
            ws.cell(row=row_num, column=10, value=entry.predefined_type)
            ws.cell(row=row_num, column=11, value=entry.material)
            ws.cell(row=row_num, column=12, value=entry.total_instance_count)
            ws.cell(row=row_num, column=13, value=entry.source_model_count)
            ws.cell(row=row_num, column=14, value=entry.confidence or '')

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Build filename
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"type_bank_{date_str}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='export-json')
    def export_json(self, request):
        """
        Export TypeBank to JSON for ML training data.

        GET /api/type-bank/export-json/
        GET /api/type-bank/export-json/?mapping_status=mapped - Only labeled types

        Returns JSON array optimized for ML:
        [
            {
                "identity": {
                    "ifc_class": "IfcWallType",
                    "type_name": "Wall 200mm",
                    "predefined_type": "STANDARD",
                    "material": "Concrete"
                },
                "labels": {
                    "ns3451_code": "222",
                    "ns3451_name": "Innervegger",
                    "discipline": "ARK",
                    "canonical_name": "Inner concrete wall",
                    "representative_unit": "m2"
                },
                "stats": {
                    "total_instance_count": 450,
                    "pct_is_external": 0.0,
                    "pct_load_bearing": 0.65,
                    "source_model_count": 12
                }
            },
            ...
        ]
        """
        # Apply filters
        queryset = self.filter_queryset(
            TypeBankEntry.objects.select_related('ns3451').all()
        )

        data = []
        for entry in queryset:
            data.append({
                'identity': {
                    'ifc_class': entry.ifc_class,
                    'type_name': entry.type_name,
                    'predefined_type': entry.predefined_type,
                    'material': entry.material,
                },
                'labels': {
                    'ns3451_code': entry.ns3451_code,
                    'ns3451_name': entry.ns3451.name if entry.ns3451 else None,
                    'discipline': entry.discipline,
                    'canonical_name': entry.canonical_name,
                    'representative_unit': entry.representative_unit,
                },
                'stats': {
                    'total_instance_count': entry.total_instance_count,
                    'pct_is_external': entry.pct_is_external,
                    'pct_load_bearing': entry.pct_load_bearing,
                    'pct_fire_rated': entry.pct_fire_rated,
                    'source_model_count': entry.source_model_count,
                },
                'meta': {
                    'id': str(entry.id),
                    'mapping_status': entry.mapping_status,
                    'confidence': entry.confidence,
                }
            })

        return Response(data)

    @action(detail=True, methods=['post'], url_path='add-alias')
    def add_alias(self, request, pk=None):
        """
        Add an alias (naming variation) to this TypeBankEntry.

        POST /api/type-bank/{id}/add-alias/
        Body: {
            "alias_type_name": "GU13",
            "alias_ifc_class": "IfcBuildingElementProxy",  # optional
            "alias_source": "Model ABC"  # optional
        }
        """
        entry = self.get_object()

        alias_type_name = request.data.get('alias_type_name')
        if not alias_type_name:
            return Response(
                {'error': 'alias_type_name is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        alias = TypeBankAlias.objects.create(
            canonical=entry,
            alias_type_name=alias_type_name,
            alias_ifc_class=request.data.get('alias_ifc_class'),
            alias_source=request.data.get('alias_source'),
        )

        return Response(TypeBankAliasSerializer(alias).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='merge-into')
    def merge_into(self, request, pk=None):
        """
        Merge another TypeBankEntry into this one.

        POST /api/type-bank/{id}/merge-into/
        Body: {
            "source_entry_id": "uuid"
        }

        This will:
        1. Move all observations from source to this entry
        2. Create an alias from source's type_name
        3. Update stats (instance counts, model counts)
        4. Delete the source entry
        """
        target = self.get_object()
        source_id = request.data.get('source_entry_id')

        if not source_id:
            return Response(
                {'error': 'source_entry_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            source = TypeBankEntry.objects.get(id=source_id)
        except TypeBankEntry.DoesNotExist:
            return Response(
                {'error': 'Source entry not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        if source.id == target.id:
            return Response(
                {'error': 'Cannot merge entry into itself'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Move observations
        observations_moved = source.observations.update(type_bank_entry=target)

        # Create alias from source's type_name if different
        if source.type_name != target.type_name:
            TypeBankAlias.objects.create(
                canonical=target,
                alias_type_name=source.type_name,
                alias_ifc_class=source.ifc_class,
                alias_source=f"Merged from {source.id}",
            )

        # Move existing aliases
        source.aliases.update(canonical=target)

        # Update stats
        target.total_instance_count += source.total_instance_count
        target.source_model_count = target.observations.values('source_model').distinct().count()
        target.save()

        # Delete source
        source_id_str = str(source.id)
        source.delete()

        return Response({
            'success': True,
            'target_id': str(target.id),
            'source_id_deleted': source_id_str,
            'observations_moved': observations_moved,
        })

    # =========================================================================
    # SEMANTIC TYPE ACTIONS (PA0802/IFC Normalization)
    # =========================================================================

    @action(detail=False, methods=['post'], url_path='auto-normalize')
    def auto_normalize(self, request):
        """
        Bulk auto-normalize unclassified TypeBankEntries.

        POST /api/type-bank/auto-normalize/
        Body (optional): {
            "overwrite": false,  // Overwrite existing semantic_type assignments
            "ifc_class": "IfcBeamType"  // Only normalize this IFC class
        }

        Returns:
        {
            "normalized": 150,
            "skipped": 20
        }
        """
        overwrite = request.data.get('overwrite', False)
        ifc_class = request.data.get('ifc_class')

        queryset = TypeBankEntry.objects.all()
        if ifc_class:
            queryset = queryset.filter(ifc_class=ifc_class)

        normalizer = get_normalizer()
        stats = normalizer.bulk_normalize(queryset, overwrite=overwrite)

        return Response(stats)

    @action(detail=True, methods=['post'], url_path='set-semantic-type')
    def set_semantic_type(self, request, pk=None):
        """
        Manually set semantic type for a TypeBankEntry.

        POST /api/type-bank/{id}/set-semantic-type/
        Body: {
            "semantic_type_code": "AB"  // PA0802 code
        }
        """
        entry = self.get_object()
        semantic_type_code = request.data.get('semantic_type_code')

        if not semantic_type_code:
            return Response(
                {'error': 'semantic_type_code is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            semantic_type = SemanticType.objects.get(code=semantic_type_code)
            entry.semantic_type = semantic_type
            entry.semantic_type_source = 'manual'
            entry.semantic_type_confidence = 1.0
            entry.save(update_fields=['semantic_type', 'semantic_type_source', 'semantic_type_confidence'])
            return Response({
                'status': 'ok',
                'semantic_type_code': semantic_type.code,
                'semantic_type_name': semantic_type.name_en,
            })
        except SemanticType.DoesNotExist:
            return Response(
                {'error': f'Invalid semantic type code: {semantic_type_code}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'], url_path='verify-semantic-type')
    def verify_semantic_type(self, request, pk=None):
        """
        Verify/confirm the current semantic type assignment.

        POST /api/type-bank/{id}/verify-semantic-type/

        Changes source to 'verified' and confidence to 1.0.
        """
        entry = self.get_object()

        if not entry.semantic_type:
            return Response(
                {'error': 'No semantic type to verify'},
                status=status.HTTP_400_BAD_REQUEST
            )

        entry.semantic_type_source = 'verified'
        entry.semantic_type_confidence = 1.0
        entry.save(update_fields=['semantic_type_source', 'semantic_type_confidence'])

        return Response({
            'status': 'verified',
            'semantic_type_code': entry.semantic_type.code,
            'semantic_type_name': entry.semantic_type.name_en,
        })

    @action(detail=True, methods=['get'], url_path='suggest-semantic-types')
    def suggest_semantic_types(self, request, pk=None):
        """
        Get suggested semantic types for a TypeBankEntry.

        GET /api/type-bank/{id}/suggest-semantic-types/

        Returns list of suggestions with confidence scores.
        """
        entry = self.get_object()
        normalizer = get_normalizer()

        suggestions = normalizer.suggest_semantic_type(
            ifc_class=entry.ifc_class,
            type_name=entry.type_name or '',
            predefined_type=entry.predefined_type or ''
        )

        # Convert SemanticType objects to serializable dicts
        result = []
        for s in suggestions:
            result.append({
                'code': s['code'],
                'name_en': s['name_en'],
                'source': s['source'],
                'confidence': s['confidence'],
                'is_primary': s['is_primary'],
                'is_common_misuse': s['is_common_misuse'],
                'note': s['note'],
            })

        return Response(result)

    @action(detail=False, methods=['get'], url_path='semantic-summary')
    def semantic_summary(self, request):
        """
        Get semantic type assignment summary.

        GET /api/type-bank/semantic-summary/

        Returns:
        {
            "total": 332,
            "with_semantic_type": 320,
            "without_semantic_type": 12,
            "by_source": {"auto_rule": 280, "auto_pattern": 30, "manual": 5, "verified": 5},
            "by_semantic_type": {"AB": 150, "AS": 50, ...}
        }
        """
        total = TypeBankEntry.objects.count()
        with_st = TypeBankEntry.objects.filter(semantic_type__isnull=False).count()
        without_st = TypeBankEntry.objects.filter(semantic_type__isnull=True).count()

        # By source
        by_source = dict(
            TypeBankEntry.objects.filter(semantic_type_source__isnull=False)
            .values('semantic_type_source')
            .annotate(count=Count('id'))
            .values_list('semantic_type_source', 'count')
        )

        # By semantic type
        by_semantic_type = dict(
            TypeBankEntry.objects.filter(semantic_type__isnull=False)
            .values('semantic_type__code', 'semantic_type__name_en')
            .annotate(count=Count('id'))
            .values_list('semantic_type__code', 'count')
        )

        return Response({
            'total': total,
            'with_semantic_type': with_st,
            'without_semantic_type': without_st,
            'coverage_percent': round((with_st / total * 100) if total > 0 else 0, 1),
            'by_source': by_source,
            'by_semantic_type': by_semantic_type,
        })


class TypeBankObservationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for TypeBank observations (where types were found).

    GET /api/type-bank-observations/?type_bank_entry={id} - Observations for an entry
    GET /api/type-bank-observations/?source_model={id} - Observations from a model
    """
    queryset = TypeBankObservation.objects.select_related(
        'type_bank_entry', 'source_model', 'source_model__project', 'source_type'
    ).all()
    serializer_class = TypeBankObservationSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['type_bank_entry', 'source_model']
    ordering = ['-observed_at']


class TypeBankAliasViewSet(viewsets.ModelViewSet):
    """
    API endpoint for TypeBank aliases (naming variations).

    GET /api/type-bank-aliases/?canonical={id} - Aliases for an entry
    POST /api/type-bank-aliases/ - Create alias
    DELETE /api/type-bank-aliases/{id}/ - Delete alias
    """
    queryset = TypeBankAlias.objects.select_related('canonical').all()
    serializer_class = TypeBankAliasSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['canonical']
    search_fields = ['alias_type_name']


# =============================================================================
# MATERIAL & PRODUCT LIBRARY VIEWSETS
# =============================================================================

class MaterialLibraryViewSet(viewsets.ModelViewSet):
    """
    API endpoint for MaterialLibrary (global canonical materials).

    MaterialLibrary contains Enova EPD material categories with:
    - Physical properties (density, thermal conductivity)
    - EPD data (GWP, Reduzer ProductID)
    - Unit of measurement (m³, m², m, kg)

    List endpoints:
    GET /api/material-library/ - List all materials (paginated)
    GET /api/material-library/?category=gypsum_standard - Filter by category
    GET /api/material-library/?source=enova - Filter by source
    GET /api/material-library/?search=concrete - Search by name

    Detail endpoints:
    GET /api/material-library/{id}/ - Get material details
    POST /api/material-library/ - Create material
    PATCH /api/material-library/{id}/ - Update material
    DELETE /api/material-library/{id}/ - Delete material

    Actions:
    GET /api/material-library/categories/ - List available categories
    GET /api/material-library/summary/ - Get summary statistics
    """

    queryset = MaterialLibrary.objects.all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'source', 'unit', 'reused_status']
    search_fields = ['name', 'category', 'description', 'manufacturer', 'product_name']
    ordering_fields = ['name', 'category', 'gwp_a1_a3', 'density_kg_m3', 'created_at']
    ordering = ['category', 'name']

    def get_serializer_class(self):
        if self.action == 'list':
            return MaterialLibraryListSerializer
        return MaterialLibrarySerializer

    @action(detail=False, methods=['get'], url_path='categories')
    def categories(self, request):
        """
        List all material categories with counts.

        GET /api/material-library/categories/

        Returns list of {category, category_display, count} for populated categories.
        """
        from django.db.models import Count

        categories = MaterialLibrary.objects.values('category').annotate(
            count=Count('id')
        ).order_by('category')

        # Add display names
        category_choices = dict(MaterialLibrary.CATEGORY_CHOICES)
        result = []
        for cat in categories:
            result.append({
                'category': cat['category'],
                'category_display': category_choices.get(cat['category'], cat['category']),
                'count': cat['count'],
            })

        # Also include empty categories (for reference)
        populated_categories = {cat['category'] for cat in categories}
        for code, display in MaterialLibrary.CATEGORY_CHOICES:
            if code not in populated_categories:
                result.append({
                    'category': code,
                    'category_display': display,
                    'count': 0,
                })

        return Response(sorted(result, key=lambda x: x['category']))

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        Get MaterialLibrary summary statistics.

        GET /api/material-library/summary/

        Returns:
        - total: Total materials
        - by_source: Count per source
        - by_unit: Count per unit
        - with_epd: Count with Reduzer ProductID
        - with_gwp: Count with GWP data
        """
        total = MaterialLibrary.objects.count()

        by_source = dict(
            MaterialLibrary.objects.filter(source__isnull=False)
            .exclude(source='')
            .values('source')
            .annotate(count=Count('id'))
            .values_list('source', 'count')
        )

        by_unit = dict(
            MaterialLibrary.objects.values('unit')
            .annotate(count=Count('id'))
            .values_list('unit', 'count')
        )

        with_epd = MaterialLibrary.objects.filter(
            reduzer_product_id__isnull=False
        ).exclude(reduzer_product_id='').count()

        with_gwp = MaterialLibrary.objects.filter(
            gwp_a1_a3__isnull=False
        ).count()

        return Response({
            'total': total,
            'by_source': by_source,
            'by_unit': by_unit,
            'with_epd': with_epd,
            'with_gwp': with_gwp,
            'epd_coverage_percent': round((with_epd / total * 100) if total > 0 else 0, 1),
        })


class ProductLibraryViewSet(viewsets.ModelViewSet):
    """
    API endpoint for ProductLibrary (discrete components).

    ProductLibrary contains products (windows, doors, fixtures) that are:
    - Counted in pieces (pcs/stk), not quantity (m³, m², kg)
    - Either homogeneous (single material) or composite (multiple materials)
    - Linked to manufacturer specs, dimensions, datasheets

    List endpoints:
    GET /api/product-library/ - List all products (paginated)
    GET /api/product-library/?category=window - Filter by category
    GET /api/product-library/?manufacturer=Velux - Filter by manufacturer
    GET /api/product-library/?is_composite=true - Filter composite products
    GET /api/product-library/?search=skylight - Search by name

    Detail endpoints:
    GET /api/product-library/{id}/ - Get product details with compositions
    POST /api/product-library/ - Create product
    PATCH /api/product-library/{id}/ - Update product
    DELETE /api/product-library/{id}/ - Delete product

    Composition actions:
    GET /api/product-library/{id}/compositions/ - Get material composition
    POST /api/product-library/{id}/compositions/ - Add material to composition
    POST /api/product-library/{id}/set-compositions/ - Replace all compositions
    """

    queryset = ProductLibrary.objects.annotate(
        _composition_count=Count('compositions')
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'manufacturer', 'is_composite', 'material_category', 'reused_status']
    search_fields = ['name', 'category', 'manufacturer', 'product_code', 'description']
    ordering_fields = ['name', 'category', 'manufacturer', 'created_at']
    ordering = ['category', 'name']

    def get_serializer_class(self):
        if self.action == 'list':
            return ProductLibraryListSerializer
        return ProductLibrarySerializer

    @action(detail=True, methods=['get'], url_path='compositions')
    def compositions(self, request, pk=None):
        """
        Get material compositions for a product.

        GET /api/product-library/{id}/compositions/

        Returns list of ProductComposition entries for this product.
        """
        product = self.get_object()
        compositions = ProductComposition.objects.filter(
            product=product
        ).select_related('material').order_by('layer_order')

        serializer = ProductCompositionSerializer(compositions, many=True)
        return Response({
            'product_id': str(product.id),
            'product_name': product.name,
            'is_composite': product.is_composite,
            'compositions': serializer.data,
        })

    @action(detail=True, methods=['post'], url_path='set-compositions')
    def set_compositions(self, request, pk=None):
        """
        Replace all compositions for a product.

        POST /api/product-library/{id}/set-compositions/
        Body: {
            "compositions": [
                {"material_id": "uuid", "quantity": 1.5, "unit": "kg", "layer_order": 1},
                {"material_id": "uuid", "quantity": 0.5, "unit": "m2", "layer_order": 2}
            ]
        }

        This will:
        1. Delete existing compositions
        2. Create new compositions from provided data
        3. Auto-set is_composite=True if multiple compositions
        """
        product = self.get_object()
        compositions_data = request.data.get('compositions', [])

        # Delete existing compositions
        ProductComposition.objects.filter(product=product).delete()

        # Create new compositions
        created = []
        for i, comp_data in enumerate(compositions_data):
            material_id = comp_data.get('material_id')
            if not material_id:
                continue

            try:
                material = MaterialLibrary.objects.get(id=material_id)
            except MaterialLibrary.DoesNotExist:
                continue

            comp = ProductComposition.objects.create(
                product=product,
                material=material,
                quantity=comp_data.get('quantity', 1.0),
                unit=comp_data.get('unit', 'kg'),
                layer_order=comp_data.get('layer_order', i + 1),
                notes=comp_data.get('notes', ''),
            )
            created.append(comp)

        # Update is_composite flag
        product.is_composite = len(created) > 1
        product.save(update_fields=['is_composite'])

        serializer = ProductCompositionSerializer(created, many=True)
        return Response({
            'product_id': str(product.id),
            'product_name': product.name,
            'is_composite': product.is_composite,
            'compositions': serializer.data,
            'created_count': len(created),
        })

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        Get ProductLibrary summary statistics.

        GET /api/product-library/summary/

        Returns:
        - total: Total products
        - by_category: Count per category
        - composite: Count of composite products
        - with_compositions: Count with material compositions defined
        """
        total = ProductLibrary.objects.count()

        by_category = dict(
            ProductLibrary.objects.filter(category__isnull=False)
            .exclude(category='')
            .values('category')
            .annotate(count=Count('id'))
            .values_list('category', 'count')
        )

        composite = ProductLibrary.objects.filter(is_composite=True).count()

        with_compositions = ProductLibrary.objects.annotate(
            comp_count=Count('compositions')
        ).filter(comp_count__gt=0).count()

        return Response({
            'total': total,
            'by_category': by_category,
            'composite': composite,
            'homogeneous': total - composite,
            'with_compositions': with_compositions,
        })


class ProductCompositionViewSet(viewsets.ModelViewSet):
    """
    API endpoint for ProductComposition (product → material links).

    GET /api/product-compositions/?product={id} - Get compositions for a product
    POST /api/product-compositions/ - Add material to product
    PATCH /api/product-compositions/{id}/ - Update composition
    DELETE /api/product-compositions/{id}/ - Remove material from product
    """

    queryset = ProductComposition.objects.select_related('product', 'material').all()
    serializer_class = ProductCompositionSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['product', 'material']
    ordering = ['product', 'layer_order']
