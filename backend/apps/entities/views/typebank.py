"""
ViewSets for the TypeBank global type classification system.

TypeBank replaces per-model TypeMapping with a shared, cross-project type library.
Types are identified by signature tuple (ifc_class, type_name, predefined_type, material).
"""
from datetime import datetime
from io import BytesIO
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Count, Prefetch
from django_filters.rest_framework import DjangoFilterBackend
from ..models import (
    TypeBankEntry, TypeBankObservation, TypeBankAlias, SemanticType,
)
from ..serializers import (
    TypeBankEntrySerializer, TypeBankEntryListSerializer, TypeBankEntryUpdateSerializer,
    TypeBankObservationSerializer, TypeBankAliasSerializer,
)
from ..services.semantic_normalizer import get_normalizer


class TypeBankEntryViewSet(viewsets.ModelViewSet):
    """
    API endpoint for TypeBank entries (global type classification system).

    TypeBank is a shared type library that:
    - Classifies types based on (ifc_class, type_name, predefined_type, material)
    - Tracks where types have been observed across models
    - Stores expert labels (NS3451, discipline, canonical name)
    - REPLACES the per-model TypeMapping system

    List endpoints:
    GET /api/type-bank/ - List all types (paginated, filterable)
    GET /api/type-bank/?ifc_class=IfcWallType - Filter by IFC class
    GET /api/type-bank/?mapping_status=pending - Filter by mapping status
    GET /api/type-bank/?ns3451_code=222 - Filter by NS3451 code
    GET /api/type-bank/?search=concrete - Search by type_name or canonical_name

    Detail endpoints:
    GET /api/type-bank/{id}/ - Get full entry with observations and aliases
    PATCH /api/type-bank/{id}/ - Update classification labels

    Actions:
    GET /api/type-bank/summary/ - Get mapping progress stats
    GET /api/type-bank/export-excel/ - Export to Excel template
    GET /api/type-bank/export-json/ - Export to JSON for ML training
    POST /api/type-bank/{id}/add-alias/ - Add alias to entry
    POST /api/type-bank/{id}/merge-into/ - Merge another entry into this one
    """

    queryset = TypeBankEntry.objects.select_related('ns3451').annotate(
        _observation_count=Count('observations')
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['ifc_class', 'mapping_status', 'ns3451_code', 'discipline', 'confidence']
    search_fields = ['type_name', 'canonical_name', 'material', 'ifc_class']
    ordering_fields = ['ifc_class', 'type_name', 'total_instance_count', 'source_model_count', 'updated_at']
    ordering = ['ifc_class', 'type_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return TypeBankEntryListSerializer
        elif self.action in ['update', 'partial_update']:
            return TypeBankEntryUpdateSerializer
        return TypeBankEntrySerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        # For detail views, prefetch observations and aliases
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related(
                Prefetch('observations', queryset=TypeBankObservation.objects.select_related(
                    'source_model', 'source_model__project', 'source_type'
                ).order_by('-observed_at')),
                'aliases'
            )
        return queryset

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        Get TypeBank mapping progress summary.

        GET /api/type-bank/summary/

        Returns:
        {
            "total": 1234,
            "mapped": 890,
            "pending": 300,
            "ignored": 20,
            "review": 24,
            "progress_percent": 72.1,
            "by_ifc_class": {"IfcWallType": 150, "IfcColumnType": 80, ...},
            "by_discipline": {"ARK": 400, "RIB": 300, ...}
        }
        """
        total = TypeBankEntry.objects.count()
        mapped = TypeBankEntry.objects.filter(mapping_status='mapped').count()
        pending = TypeBankEntry.objects.filter(mapping_status='pending').count()
        ignored = TypeBankEntry.objects.filter(mapping_status='ignored').count()
        review = TypeBankEntry.objects.filter(mapping_status='review').count()

        # By IFC class
        by_class = dict(
            TypeBankEntry.objects.values('ifc_class')
            .annotate(count=Count('id'))
            .values_list('ifc_class', 'count')
        )

        # By discipline (only for mapped entries)
        by_discipline = dict(
            TypeBankEntry.objects.filter(discipline__isnull=False)
            .exclude(discipline='')
            .values('discipline')
            .annotate(count=Count('id'))
            .values_list('discipline', 'count')
        )

        return Response({
            'total': total,
            'mapped': mapped,
            'pending': pending,
            'ignored': ignored,
            'review': review,
            'progress_percent': round((mapped / total * 100) if total > 0 else 0, 1),
            'by_ifc_class': by_class,
            'by_discipline': by_discipline,
        })

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Export TypeBank to Excel for batch editing.

        GET /api/type-bank/export-excel/
        GET /api/type-bank/export-excel/?mapping_status=pending - Only pending types
        GET /api/type-bank/export-excel/?ifc_class=IfcWallType - Filter by class

        Returns Excel file with editable columns:
        - NS3451 Code, Discipline, Canonical Name, Representative Unit, Status, Notes
        And read-only columns:
        - IFC Class, Type Name, Predefined Type, Material, Instance Count, Model Count
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # Apply filters
        queryset = self.filter_queryset(self.get_queryset())

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TypeBank"

        # Headers
        headers = [
            # Editable columns (A-F)
            ('NS3451 Code', 12),
            ('Discipline', 10),
            ('Canonical Name', 30),
            ('Unit', 8),
            ('Status', 12),
            ('Notes', 40),
            # Read-only columns (G-N)
            ('ID', 36),
            ('IFC Class', 20),
            ('Type Name', 40),
            ('Predefined Type', 15),
            ('Material', 30),
            ('Instance Count', 14),
            ('Model Count', 12),
            ('Confidence', 12),
        ]

        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        readonly_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        for col, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[cell.column_letter].width = width

        # Data rows
        for row_num, entry in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=entry.ns3451_code or '')
            ws.cell(row=row_num, column=2, value=entry.discipline or '')
            ws.cell(row=row_num, column=3, value=entry.canonical_name or '')
            ws.cell(row=row_num, column=4, value=entry.representative_unit or '')
            ws.cell(row=row_num, column=5, value=entry.mapping_status)
            ws.cell(row=row_num, column=6, value=entry.notes or '')

            # Read-only columns (grey background)
            for col in range(7, 15):
                ws.cell(row=row_num, column=col).fill = readonly_fill

            ws.cell(row=row_num, column=7, value=str(entry.id))
            ws.cell(row=row_num, column=8, value=entry.ifc_class)
            ws.cell(row=row_num, column=9, value=entry.type_name)
            ws.cell(row=row_num, column=10, value=entry.predefined_type)
            ws.cell(row=row_num, column=11, value=entry.material)
            ws.cell(row=row_num, column=12, value=entry.total_instance_count)
            ws.cell(row=row_num, column=13, value=entry.source_model_count)
            ws.cell(row=row_num, column=14, value=entry.confidence or '')

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Build filename
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"type_bank_{date_str}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='export-json')
    def export_json(self, request):
        """
        Export TypeBank to JSON for ML training data.

        GET /api/type-bank/export-json/
        GET /api/type-bank/export-json/?mapping_status=mapped - Only labeled types

        Returns JSON array optimized for ML:
        [
            {
                "identity": {
                    "ifc_class": "IfcWallType",
                    "type_name": "Wall 200mm",
                    "predefined_type": "STANDARD",
                    "material": "Concrete"
                },
                "labels": {
                    "ns3451_code": "222",
                    "ns3451_name": "Innervegger",
                    "discipline": "ARK",
                    "canonical_name": "Inner concrete wall",
                    "representative_unit": "m2"
                },
                "stats": {
                    "total_instance_count": 450,
                    "pct_is_external": 0.0,
                    "pct_load_bearing": 0.65,
                    "source_model_count": 12
                }
            },
            ...
        ]
        """
        # Apply filters
        queryset = self.filter_queryset(
            TypeBankEntry.objects.select_related('ns3451').all()
        )

        data = []
        for entry in queryset:
            data.append({
                'identity': {
                    'ifc_class': entry.ifc_class,
                    'type_name': entry.type_name,
                    'predefined_type': entry.predefined_type,
                    'material': entry.material,
                },
                'labels': {
                    'ns3451_code': entry.ns3451_code,
                    'ns3451_name': entry.ns3451.name if entry.ns3451 else None,
                    'discipline': entry.discipline,
                    'canonical_name': entry.canonical_name,
                    'representative_unit': entry.representative_unit,
                },
                'stats': {
                    'total_instance_count': entry.total_instance_count,
                    'pct_is_external': entry.pct_is_external,
                    'pct_load_bearing': entry.pct_load_bearing,
                    'pct_fire_rated': entry.pct_fire_rated,
                    'source_model_count': entry.source_model_count,
                },
                'meta': {
                    'id': str(entry.id),
                    'mapping_status': entry.mapping_status,
                    'confidence': entry.confidence,
                }
            })

        return Response(data)

    @action(detail=True, methods=['post'], url_path='add-alias')
    def add_alias(self, request, pk=None):
        """
        Add an alias (naming variation) to this TypeBankEntry.

        POST /api/type-bank/{id}/add-alias/
        Body: {
            "alias_type_name": "GU13",
            "alias_ifc_class": "IfcBuildingElementProxy",  # optional
            "alias_source": "Model ABC"  # optional
        }
        """
        entry = self.get_object()

        alias_type_name = request.data.get('alias_type_name')
        if not alias_type_name:
            return Response(
                {'error': 'alias_type_name is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        alias = TypeBankAlias.objects.create(
            canonical=entry,
            alias_type_name=alias_type_name,
            alias_ifc_class=request.data.get('alias_ifc_class'),
            alias_source=request.data.get('alias_source'),
        )

        return Response(TypeBankAliasSerializer(alias).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='merge-into')
    def merge_into(self, request, pk=None):
        """
        Merge another TypeBankEntry into this one.

        POST /api/type-bank/{id}/merge-into/
        Body: {
            "source_entry_id": "uuid"
        }

        This will:
        1. Move all observations from source to this entry
        2. Create an alias from source's type_name
        3. Update stats (instance counts, model counts)
        4. Delete the source entry
        """
        target = self.get_object()
        source_id = request.data.get('source_entry_id')

        if not source_id:
            return Response(
                {'error': 'source_entry_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            source = TypeBankEntry.objects.get(id=source_id)
        except TypeBankEntry.DoesNotExist:
            return Response(
                {'error': 'Source entry not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        if source.id == target.id:
            return Response(
                {'error': 'Cannot merge entry into itself'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Move observations
        observations_moved = source.observations.update(type_bank_entry=target)

        # Create alias from source's type_name if different
        if source.type_name != target.type_name:
            TypeBankAlias.objects.create(
                canonical=target,
                alias_type_name=source.type_name,
                alias_ifc_class=source.ifc_class,
                alias_source=f"Merged from {source.id}",
            )

        # Move existing aliases
        source.aliases.update(canonical=target)

        # Update stats
        target.total_instance_count += source.total_instance_count
        target.source_model_count = target.observations.values('source_model').distinct().count()
        target.save()

        # Delete source
        source_id_str = str(source.id)
        source.delete()

        return Response({
            'success': True,
            'target_id': str(target.id),
            'source_id_deleted': source_id_str,
            'observations_moved': observations_moved,
        })

    # =========================================================================
    # SEMANTIC TYPE ACTIONS (PA0802/IFC Normalization)
    # =========================================================================

    @action(detail=False, methods=['post'], url_path='auto-normalize')
    def auto_normalize(self, request):
        """
        Bulk auto-normalize unclassified TypeBankEntries.

        POST /api/type-bank/auto-normalize/
        Body (optional): {
            "overwrite": false,  // Overwrite existing semantic_type assignments
            "ifc_class": "IfcBeamType"  // Only normalize this IFC class
        }

        Returns:
        {
            "normalized": 150,
            "skipped": 20
        }
        """
        overwrite = request.data.get('overwrite', False)
        ifc_class = request.data.get('ifc_class')

        queryset = TypeBankEntry.objects.all()
        if ifc_class:
            queryset = queryset.filter(ifc_class=ifc_class)

        normalizer = get_normalizer()
        stats = normalizer.bulk_normalize(queryset, overwrite=overwrite)

        return Response(stats)

    @action(detail=True, methods=['post'], url_path='set-semantic-type')
    def set_semantic_type(self, request, pk=None):
        """
        Manually set semantic type for a TypeBankEntry.

        POST /api/type-bank/{id}/set-semantic-type/
        Body: {
            "semantic_type_code": "AB"  // PA0802 code
        }
        """
        entry = self.get_object()
        semantic_type_code = request.data.get('semantic_type_code')

        if not semantic_type_code:
            return Response(
                {'error': 'semantic_type_code is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            semantic_type = SemanticType.objects.get(code=semantic_type_code)
            entry.semantic_type = semantic_type
            entry.semantic_type_source = 'manual'
            entry.semantic_type_confidence = 1.0
            entry.save(update_fields=['semantic_type', 'semantic_type_source', 'semantic_type_confidence'])
            return Response({
                'status': 'ok',
                'semantic_type_code': semantic_type.code,
                'semantic_type_name': semantic_type.name_en,
            })
        except SemanticType.DoesNotExist:
            return Response(
                {'error': f'Invalid semantic type code: {semantic_type_code}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'], url_path='verify-semantic-type')
    def verify_semantic_type(self, request, pk=None):
        """
        Verify/confirm the current semantic type assignment.

        POST /api/type-bank/{id}/verify-semantic-type/

        Changes source to 'verified' and confidence to 1.0.
        """
        entry = self.get_object()

        if not entry.semantic_type:
            return Response(
                {'error': 'No semantic type to verify'},
                status=status.HTTP_400_BAD_REQUEST
            )

        entry.semantic_type_source = 'verified'
        entry.semantic_type_confidence = 1.0
        entry.save(update_fields=['semantic_type_source', 'semantic_type_confidence'])

        return Response({
            'status': 'verified',
            'semantic_type_code': entry.semantic_type.code,
            'semantic_type_name': entry.semantic_type.name_en,
        })

    @action(detail=True, methods=['get'], url_path='suggest-semantic-types')
    def suggest_semantic_types(self, request, pk=None):
        """
        Get suggested semantic types for a TypeBankEntry.

        GET /api/type-bank/{id}/suggest-semantic-types/

        Returns list of suggestions with confidence scores.
        """
        entry = self.get_object()
        normalizer = get_normalizer()

        suggestions = normalizer.suggest_semantic_type(
            ifc_class=entry.ifc_class,
            type_name=entry.type_name or '',
            predefined_type=entry.predefined_type or ''
        )

        # Convert SemanticType objects to serializable dicts
        result = []
        for s in suggestions:
            result.append({
                'code': s['code'],
                'name_en': s['name_en'],
                'source': s['source'],
                'confidence': s['confidence'],
                'is_primary': s['is_primary'],
                'is_common_misuse': s['is_common_misuse'],
                'note': s['note'],
            })

        return Response(result)

    @action(detail=False, methods=['get'], url_path='semantic-summary')
    def semantic_summary(self, request):
        """
        Get semantic type assignment summary.

        GET /api/type-bank/semantic-summary/

        Returns:
        {
            "total": 332,
            "with_semantic_type": 320,
            "without_semantic_type": 12,
            "by_source": {"auto_rule": 280, "auto_pattern": 30, "manual": 5, "verified": 5},
            "by_semantic_type": {"AB": 150, "AS": 50, ...}
        }
        """
        total = TypeBankEntry.objects.count()
        with_st = TypeBankEntry.objects.filter(semantic_type__isnull=False).count()
        without_st = TypeBankEntry.objects.filter(semantic_type__isnull=True).count()

        # By source
        by_source = dict(
            TypeBankEntry.objects.filter(semantic_type_source__isnull=False)
            .values('semantic_type_source')
            .annotate(count=Count('id'))
            .values_list('semantic_type_source', 'count')
        )

        # By semantic type
        by_semantic_type = dict(
            TypeBankEntry.objects.filter(semantic_type__isnull=False)
            .values('semantic_type__code', 'semantic_type__name_en')
            .annotate(count=Count('id'))
            .values_list('semantic_type__code', 'count')
        )

        return Response({
            'total': total,
            'with_semantic_type': with_st,
            'without_semantic_type': without_st,
            'coverage_percent': round((with_st / total * 100) if total > 0 else 0, 1),
            'by_source': by_source,
            'by_semantic_type': by_semantic_type,
        })


class TypeBankObservationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for TypeBank observations (where types were found).

    GET /api/type-bank-observations/?type_bank_entry={id} - Observations for an entry
    GET /api/type-bank-observations/?source_model={id} - Observations from a model
    """
    queryset = TypeBankObservation.objects.select_related(
        'type_bank_entry', 'source_model', 'source_model__project', 'source_type'
    ).all()
    serializer_class = TypeBankObservationSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['type_bank_entry', 'source_model']
    ordering = ['-observed_at']


class TypeBankAliasViewSet(viewsets.ModelViewSet):
    """
    API endpoint for TypeBank aliases (naming variations).

    GET /api/type-bank-aliases/?canonical={id} - Aliases for an entry
    POST /api/type-bank-aliases/ - Create alias
    DELETE /api/type-bank-aliases/{id}/ - Delete alias
    """
    queryset = TypeBankAlias.objects.select_related('canonical').all()
    serializer_class = TypeBankAliasSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['canonical']
    search_fields = ['alias_type_name']
