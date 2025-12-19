"""
Reduzer export service for type mapping.

Generates Excel file compatible with Reduzer import format.
Columns:
- description: Type name/description
- NS3451:2009: NS3451 building classification code
- quantity: Aggregated quantity (sum of instances)
- unit: Unit of measurement (kg, m, m2, m3, stk)
- component: Grouping (discipline or NS3451 category)
- productIDType: Type of product ID (EPD, NOBB, etc.)
- productID: Product/EPD identifier
- notes: Additional comments

Reference template: reduzer_import_template.xlsx
"""
from typing import Dict, List, Any
from io import BytesIO
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count

from ..models import (
    IFCType,
    TypeAssignment,
    TypeMapping,
    TypeDefinitionLayer,
    IFCEntity,
)


# Reduzer-compatible columns
REDUZER_COLUMNS = [
    ('description', 40),
    ('NS3451:2009', 15),
    ('quantity', 12),
    ('unit', 8),
    ('component', 20),
    ('productIDType', 15),
    ('productID', 20),
    ('notes', 40),
]

# Unit mapping from our internal format to Reduzer format
UNIT_MAP = {
    'pcs': 'stk',  # pieces → stykk (Norwegian)
    'm': 'm',
    'm2': 'm2',
    'm3': 'm3',
    None: 'stk',  # default
    '': 'stk',
}

# Discipline inference from IFC type
DISCIPLINE_MAP = {
    'IfcWall': 'ARK',
    'IfcWallType': 'ARK',
    'IfcDoor': 'ARK',
    'IfcDoorType': 'ARK',
    'IfcWindow': 'ARK',
    'IfcWindowType': 'ARK',
    'IfcSlab': 'ARK',
    'IfcSlabType': 'ARK',
    'IfcRoof': 'ARK',
    'IfcRoofType': 'ARK',
    'IfcStair': 'ARK',
    'IfcStairType': 'ARK',
    'IfcRailing': 'ARK',
    'IfcRailingType': 'ARK',
    'IfcCurtainWall': 'ARK',
    'IfcCurtainWallType': 'ARK',
    'IfcColumn': 'RIB',
    'IfcColumnType': 'RIB',
    'IfcBeam': 'RIB',
    'IfcBeamType': 'RIB',
    'IfcFooting': 'RIB',
    'IfcFootingType': 'RIB',
    'IfcPile': 'RIB',
    'IfcPileType': 'RIB',
    'IfcPlate': 'RIB',
    'IfcPlateType': 'RIB',
    'IfcMember': 'RIB',
    'IfcMemberType': 'RIB',
    'IfcPipeFitting': 'RIV',
    'IfcPipeFittingType': 'RIV',
    'IfcPipeSegment': 'RIV',
    'IfcPipeSegmentType': 'RIV',
    'IfcValve': 'RIV',
    'IfcValveType': 'RIV',
    'IfcPump': 'RIV',
    'IfcPumpType': 'RIV',
    'IfcTank': 'RIV',
    'IfcTankType': 'RIV',
    'IfcSanitaryTerminal': 'RIV',
    'IfcSanitaryTerminalType': 'RIV',
    'IfcCableCarrierFitting': 'RIE',
    'IfcCableCarrierFittingType': 'RIE',
    'IfcCableCarrierSegment': 'RIE',
    'IfcCableCarrierSegmentType': 'RIE',
    'IfcCableFitting': 'RIE',
    'IfcCableFittingType': 'RIE',
    'IfcCableSegment': 'RIE',
    'IfcCableSegmentType': 'RIE',
    'IfcElectricDistributionBoard': 'RIE',
    'IfcElectricDistributionBoardType': 'RIE',
    'IfcLightFixture': 'RIE',
    'IfcLightFixtureType': 'RIE',
    'IfcOutlet': 'RIE',
    'IfcOutletType': 'RIE',
    'IfcSwitchingDevice': 'RIE',
    'IfcSwitchingDeviceType': 'RIE',
    'IfcDuctFitting': 'RIV',
    'IfcDuctFittingType': 'RIV',
    'IfcDuctSegment': 'RIV',
    'IfcDuctSegmentType': 'RIV',
    'IfcAirTerminal': 'RIV',
    'IfcAirTerminalType': 'RIV',
    'IfcFan': 'RIV',
    'IfcFanType': 'RIV',
    'IfcCoil': 'RIV',
    'IfcCoilType': 'RIV',
    'IfcBoiler': 'RIV',
    'IfcBoilerType': 'RIV',
    'IfcChiller': 'RIV',
    'IfcChillerType': 'RIV',
}


class ReduzerExporter:
    """Generates Excel file compatible with Reduzer import format."""

    def __init__(self, model_id: str, include_unmapped: bool = False):
        """
        Initialize Reduzer exporter.

        Args:
            model_id: UUID of the model to export
            include_unmapped: Whether to include types without NS3451 code
        """
        self.model_id = model_id
        self.include_unmapped = include_unmapped
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Reduzer Import"

        # Styles
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF')
        self.data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        self.border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

    def export(self) -> BytesIO:
        """Generate Reduzer-compatible Excel file and return as BytesIO."""
        # Get types with aggregated data
        types_data = self._get_types_for_reduzer()

        # Write headers
        self._write_headers()

        # Write data rows
        row_idx = 2
        for type_data in types_data:
            # Expand type into rows per material layer (if layers exist)
            rows = self._expand_to_rows(type_data)
            for row in rows:
                self._write_row(row_idx, row)
                row_idx += 1

        # Setup formatting
        self._setup_column_widths()
        self._freeze_panes()

        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def _get_types_for_reduzer(self) -> List[Dict[str, Any]]:
        """Query types with aggregated quantities for Reduzer export."""
        # Only get types that have mappings with NS3451 codes (unless include_unmapped)
        types_query = IFCType.objects.filter(model_id=self.model_id).select_related('mapping')

        if not self.include_unmapped:
            types_query = types_query.filter(
                mapping__isnull=False,
                mapping__ns3451_code__isnull=False,
            ).exclude(mapping__ns3451_code='')

        result = []
        for ifc_type in types_query:
            mapping = getattr(ifc_type, 'mapping', None)

            # Get entity IDs for this type
            entity_ids = list(
                TypeAssignment.objects
                .filter(type=ifc_type)
                .values_list('entity_id', flat=True)
            )

            if not entity_ids:
                continue

            # Get unit for quantity aggregation
            unit = mapping.representative_unit if mapping else 'pcs'

            # Aggregate quantity based on unit
            quantity = self._aggregate_quantity(entity_ids, unit)

            # Get discipline
            discipline = DISCIPLINE_MAP.get(ifc_type.ifc_type, 'ARK')
            if mapping and mapping.discipline:
                discipline = mapping.discipline

            # Get definition layers for EPD info
            layers = []
            if mapping:
                layers = list(
                    TypeDefinitionLayer.objects
                    .filter(type_mapping=mapping)
                    .order_by('layer_order')
                    .values('material_name', 'thickness_mm', 'epd_id')
                )

            # Get component name (NS3451 2-digit parent or discipline)
            component = discipline
            if mapping and mapping.ns3451_code:
                # Use first 2 digits of NS3451 as component grouping
                component = mapping.ns3451_code[:2] if len(mapping.ns3451_code) >= 2 else discipline

            result.append({
                'type_guid': ifc_type.type_guid,
                'type_name': ifc_type.type_name or ifc_type.ifc_type,
                'ifc_type': ifc_type.ifc_type,
                'ns3451_code': mapping.ns3451_code if mapping else '',
                'quantity': quantity,
                'unit': UNIT_MAP.get(unit, 'stk'),
                'component': component,
                'discipline': discipline,
                'notes': mapping.notes if mapping else '',
                'layers': layers,
                'instance_count': len(entity_ids),
            })

        return result

    def _aggregate_quantity(self, entity_ids: List, unit: str) -> float:
        """Aggregate quantity from entity instances based on unit type."""
        if not entity_ids:
            return 0

        if unit == 'pcs' or not unit:
            # Count instances
            return len(entity_ids)

        elif unit == 'm':
            # Sum length
            result = IFCEntity.objects.filter(id__in=entity_ids).aggregate(
                total=Sum('length')
            )
            return round(result['total'] or 0, 2)

        elif unit == 'm2':
            # Sum area
            result = IFCEntity.objects.filter(id__in=entity_ids).aggregate(
                total=Sum('area')
            )
            return round(result['total'] or 0, 2)

        elif unit == 'm3':
            # Sum volume
            result = IFCEntity.objects.filter(id__in=entity_ids).aggregate(
                total=Sum('volume')
            )
            return round(result['total'] or 0, 2)

        else:
            # Fallback to count
            return len(entity_ids)

    def _expand_to_rows(self, type_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Expand a type into multiple rows if it has material layers with EPD IDs.

        Reduzer expects one row per product/EPD, so a wall with 3 layers
        should become 3 rows (one per material with EPD).
        """
        layers = type_data.get('layers', [])

        # If no layers or no layers have EPD, return single row
        layers_with_epd = [l for l in layers if l.get('epd_id')]

        if not layers_with_epd:
            # Single row for the type
            return [{
                'description': type_data['type_name'],
                'ns3451': type_data['ns3451_code'],
                'quantity': type_data['quantity'],
                'unit': type_data['unit'],
                'component': type_data['component'],
                'product_type': '',
                'product_id': '',
                'notes': type_data['notes'],
            }]

        # Multiple rows - one per layer with EPD
        rows = []
        for layer in layers_with_epd:
            # Create description with layer info
            layer_desc = f"{type_data['type_name']} - {layer['material_name']}"
            if layer.get('thickness_mm'):
                layer_desc += f" ({layer['thickness_mm']}mm)"

            rows.append({
                'description': layer_desc,
                'ns3451': type_data['ns3451_code'],
                'quantity': type_data['quantity'],  # Same quantity for each layer
                'unit': type_data['unit'],
                'component': type_data['component'],
                'product_type': 'EPD',
                'product_id': layer['epd_id'],
                'notes': type_data['notes'],
            })

        return rows

    def _write_headers(self):
        """Write header row with styling."""
        for col_idx, (header, _) in enumerate(REDUZER_COLUMNS, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

    def _write_row(self, row_idx: int, data: Dict[str, Any]):
        """Write a single data row in Reduzer format."""
        row_values = [
            data.get('description', ''),
            data.get('ns3451', ''),
            data.get('quantity', 0),
            data.get('unit', 'stk'),
            data.get('component', ''),
            data.get('product_type', ''),
            data.get('product_id', ''),
            data.get('notes', ''),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = self.worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = self.data_fill
            cell.border = self.border
            cell.alignment = Alignment(vertical='center')

    def _setup_column_widths(self):
        """Set column widths."""
        for col_idx, (_, width) in enumerate(REDUZER_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            self.worksheet.column_dimensions[col_letter].width = width

    def _freeze_panes(self):
        """Freeze header row."""
        self.worksheet.freeze_panes = 'A2'


def export_types_to_reduzer(model_id: str, include_unmapped: bool = False) -> BytesIO:
    """
    Export types for a model to Reduzer-compatible Excel format.

    Args:
        model_id: UUID of the model
        include_unmapped: Whether to include types without NS3451 mapping

    Returns:
        BytesIO containing the Excel file
    """
    exporter = ReduzerExporter(model_id, include_unmapped=include_unmapped)
    return exporter.export()
