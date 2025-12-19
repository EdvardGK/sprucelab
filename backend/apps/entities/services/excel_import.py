"""
Excel import service for type mapping.

Parses uploaded Excel file and bulk-updates TypeMapping records.
"""
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook
from django.db import transaction
from django.utils import timezone

from ..models import IFCType, TypeMapping, NS3451Code


@dataclass
class ImportError:
    """Error during import."""
    row: int
    type_guid: str
    error: str


@dataclass
class ImportWarning:
    """Warning during import."""
    row: int
    type_guid: str
    warning: str


@dataclass
class ImportResult:
    """Result of Excel import operation."""
    success: bool
    total_rows: int = 0
    updated: int = 0
    created: int = 0
    skipped: int = 0
    errors: List[ImportError] = field(default_factory=list)
    warnings: List[ImportWarning] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for API response."""
        return {
            'success': self.success,
            'summary': {
                'total_rows': self.total_rows,
                'updated': self.updated,
                'created': self.created,
                'skipped': self.skipped,
                'error_count': len(self.errors),
            },
            'errors': [
                {'row': e.row, 'type_guid': e.type_guid, 'error': e.error}
                for e in self.errors
            ],
            'warnings': [
                {'row': w.row, 'type_guid': w.type_guid, 'warning': w.warning}
                for w in self.warnings
            ],
        }


# Expected column positions (0-indexed)
COL_NS3451 = 0       # A - NS3451 Code
COL_UNIT = 1         # B - Unit
COL_NOTES = 2        # C - Notes
COL_STATUS = 3       # D - Status
COL_TYPE_GUID = 17   # R - Type GUID (for matching)

VALID_UNITS = {'pcs', 'm', 'm2', 'm3', ''}
VALID_STATUSES = {'pending', 'mapped', 'ignored', 'review', 'followup'}


class TypeExcelImporter:
    """Imports type mappings from Excel file."""

    def __init__(self, model_id: str, file_content: BytesIO, username: str = None):
        self.model_id = model_id
        self.file_content = file_content
        self.username = username
        self.result = ImportResult(success=True)

        # Cache for validation
        self._valid_ns3451_codes: Optional[set] = None
        self._type_guid_to_id: Optional[Dict[str, str]] = None

    def import_file(self) -> ImportResult:
        """Parse Excel and update TypeMapping records."""
        try:
            workbook = load_workbook(self.file_content, read_only=True)
            worksheet = workbook.active

            # Load validation caches
            self._load_caches()

            # Parse rows (skip header)
            rows_to_process = []
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue  # Skip empty rows

                parsed = self._parse_row(row_idx, row)
                if parsed:
                    rows_to_process.append(parsed)

            self.result.total_rows = len(rows_to_process)

            # Bulk update in transaction
            if rows_to_process:
                self._bulk_update(rows_to_process)

            workbook.close()

        except Exception as e:
            self.result.success = False
            self.result.errors.append(ImportError(
                row=0,
                type_guid='',
                error=f"Failed to parse Excel file: {str(e)}"
            ))

        # Mark as failed if too many errors
        if len(self.result.errors) > self.result.total_rows * 0.5:
            self.result.success = False

        return self.result

    def _load_caches(self):
        """Load validation caches for NS3451 codes and type GUIDs."""
        # Load valid NS3451 codes
        self._valid_ns3451_codes = set(
            NS3451Code.objects.values_list('code', flat=True)
        )

        # Load type GUIDs for this model
        types = IFCType.objects.filter(model_id=self.model_id).values('id', 'type_guid')
        self._type_guid_to_id = {t['type_guid']: str(t['id']) for t in types}

    def _parse_row(self, row_idx: int, row: tuple) -> Optional[Dict[str, Any]]:
        """
        Parse a single row and validate.

        Returns dict with mapping data or None if row should be skipped.
        """
        # Get type GUID (required for matching)
        type_guid = self._get_cell_value(row, COL_TYPE_GUID)
        if not type_guid:
            self.result.skipped += 1
            return None

        # Validate type exists in this model
        if type_guid not in self._type_guid_to_id:
            self.result.errors.append(ImportError(
                row=row_idx,
                type_guid=type_guid,
                error=f"Type GUID not found in model"
            ))
            return None

        # Get editable values
        ns3451_code = self._get_cell_value(row, COL_NS3451)
        unit = self._get_cell_value(row, COL_UNIT)
        notes = self._get_cell_value(row, COL_NOTES)
        status = self._get_cell_value(row, COL_STATUS)

        # Validate NS3451 code
        if ns3451_code and ns3451_code not in self._valid_ns3451_codes:
            self.result.errors.append(ImportError(
                row=row_idx,
                type_guid=type_guid,
                error=f"Invalid NS3451 code: {ns3451_code}"
            ))
            return None

        # Validate unit
        if unit and unit.lower() not in VALID_UNITS:
            # Try to normalize
            unit_map = {'m²': 'm2', 'm³': 'm3', 'stk': 'pcs', 'piece': 'pcs', 'pieces': 'pcs'}
            unit = unit_map.get(unit.lower(), unit)
            if unit.lower() not in VALID_UNITS:
                self.result.errors.append(ImportError(
                    row=row_idx,
                    type_guid=type_guid,
                    error=f"Invalid unit: {unit}. Use: pcs, m, m2, m3"
                ))
                return None

        # Validate/normalize status
        if status:
            status_lower = status.lower().strip()
            if status_lower not in VALID_STATUSES:
                self.result.warnings.append(ImportWarning(
                    row=row_idx,
                    type_guid=type_guid,
                    warning=f"Invalid status '{status}', defaulting to 'pending'"
                ))
                status = 'pending'
            else:
                status = status_lower
        else:
            # Auto-set status based on NS3451 code
            status = 'mapped' if ns3451_code else 'pending'

        # Add warning if NS3451 empty but status is mapped
        if not ns3451_code and status == 'mapped':
            self.result.warnings.append(ImportWarning(
                row=row_idx,
                type_guid=type_guid,
                warning="Status is 'mapped' but NS3451 code is empty"
            ))

        return {
            'type_id': self._type_guid_to_id[type_guid],
            'type_guid': type_guid,
            'ns3451_code': ns3451_code or None,
            'representative_unit': unit.lower() if unit else None,
            'notes': notes or None,
            'mapping_status': status,
            'row_idx': row_idx,
        }

    def _get_cell_value(self, row: tuple, col_idx: int) -> str:
        """Get cell value as string, handling None and various types."""
        if col_idx >= len(row):
            return ''
        value = row[col_idx]
        if value is None:
            return ''
        return str(value).strip()

    @transaction.atomic
    def _bulk_update(self, rows: List[Dict[str, Any]]):
        """Bulk update TypeMapping records."""
        now = timezone.now()

        for row_data in rows:
            type_id = row_data['type_id']

            # Get or create TypeMapping
            mapping, created = TypeMapping.objects.get_or_create(
                ifc_type_id=type_id,
                defaults={
                    'ns3451_code': row_data['ns3451_code'],
                    'representative_unit': row_data['representative_unit'],
                    'notes': row_data['notes'],
                    'mapping_status': row_data['mapping_status'],
                    'mapped_by': self.username,
                    'mapped_at': now if row_data['ns3451_code'] else None,
                }
            )

            if created:
                self.result.created += 1
            else:
                # Update existing mapping
                changed = False

                if row_data['ns3451_code'] != mapping.ns3451_code:
                    mapping.ns3451_code = row_data['ns3451_code']
                    # Also update NS3451 FK if code is valid
                    if row_data['ns3451_code']:
                        try:
                            mapping.ns3451 = NS3451Code.objects.get(code=row_data['ns3451_code'])
                        except NS3451Code.DoesNotExist:
                            mapping.ns3451 = None
                    else:
                        mapping.ns3451 = None
                    changed = True

                if row_data['representative_unit'] != mapping.representative_unit:
                    mapping.representative_unit = row_data['representative_unit']
                    changed = True

                if row_data['notes'] != mapping.notes:
                    mapping.notes = row_data['notes']
                    changed = True

                if row_data['mapping_status'] != mapping.mapping_status:
                    mapping.mapping_status = row_data['mapping_status']
                    changed = True

                if changed:
                    mapping.mapped_by = self.username
                    if row_data['ns3451_code']:
                        mapping.mapped_at = now
                    mapping.save()
                    self.result.updated += 1
                else:
                    self.result.skipped += 1


def import_types_from_excel(
    model_id: str,
    file_content: BytesIO,
    username: str = None
) -> ImportResult:
    """
    Import type mappings from Excel file.

    Args:
        model_id: UUID of the model
        file_content: BytesIO containing the Excel file
        username: User performing the import

    Returns:
        ImportResult with success status and details
    """
    importer = TypeExcelImporter(model_id, file_content, username)
    return importer.import_file()
