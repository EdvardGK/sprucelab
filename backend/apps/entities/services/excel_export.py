"""
Excel export service for type mapping.

Generates Excel template with:
- Editable columns (A-D): NS3451 Code, Unit, Notes, Status
- Read-only columns (E-R): Type metadata and aggregated properties
- Rows grouped by IfcEntity
"""
from typing import Dict, List, Any, Optional
from collections import defaultdict
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.db.models import Count, Q

from ..models import (
    IFCType,
    TypeAssignment,
    TypeMapping,
    PropertySet,
    MaterialAssignment,
    SystemMembership,
    NS3451Code,
    IFCEntity,
)


# Discipline inference from IFC type
DISCIPLINE_MAP = {
    'IfcWall': 'ARK',
    'IfcWallType': 'ARK',
    'IfcDoor': 'ARK',
    'IfcDoorType': 'ARK',
    'IfcWindow': 'ARK',
    'IfcWindowType': 'ARK',
    'IfcSlab': 'ARK',
    'IfcSlabType': 'ARK',
    'IfcRoof': 'ARK',
    'IfcRoofType': 'ARK',
    'IfcStair': 'ARK',
    'IfcStairType': 'ARK',
    'IfcRailing': 'ARK',
    'IfcRailingType': 'ARK',
    'IfcCurtainWall': 'ARK',
    'IfcCurtainWallType': 'ARK',
    'IfcColumn': 'RIB',
    'IfcColumnType': 'RIB',
    'IfcBeam': 'RIB',
    'IfcBeamType': 'RIB',
    'IfcFooting': 'RIB',
    'IfcFootingType': 'RIB',
    'IfcPile': 'RIB',
    'IfcPileType': 'RIB',
    'IfcPlate': 'RIB',
    'IfcPlateType': 'RIB',
    'IfcMember': 'RIB',
    'IfcMemberType': 'RIB',
    'IfcPipeFitting': 'RIV',
    'IfcPipeFittingType': 'RIV',
    'IfcPipeSegment': 'RIV',
    'IfcPipeSegmentType': 'RIV',
    'IfcValve': 'RIV',
    'IfcValveType': 'RIV',
    'IfcPump': 'RIV',
    'IfcPumpType': 'RIV',
    'IfcTank': 'RIV',
    'IfcTankType': 'RIV',
    'IfcSanitaryTerminal': 'RIV',
    'IfcSanitaryTerminalType': 'RIV',
    'IfcCableCarrierFitting': 'RIE',
    'IfcCableCarrierFittingType': 'RIE',
    'IfcCableCarrierSegment': 'RIE',
    'IfcCableCarrierSegmentType': 'RIE',
    'IfcCableFitting': 'RIE',
    'IfcCableFittingType': 'RIE',
    'IfcCableSegment': 'RIE',
    'IfcCableSegmentType': 'RIE',
    'IfcElectricDistributionBoard': 'RIE',
    'IfcElectricDistributionBoardType': 'RIE',
    'IfcLightFixture': 'RIE',
    'IfcLightFixtureType': 'RIE',
    'IfcOutlet': 'RIE',
    'IfcOutletType': 'RIE',
    'IfcSwitchingDevice': 'RIE',
    'IfcSwitchingDeviceType': 'RIE',
    'IfcDuctFitting': 'RIV',
    'IfcDuctFittingType': 'RIV',
    'IfcDuctSegment': 'RIV',
    'IfcDuctSegmentType': 'RIV',
    'IfcAirTerminal': 'RIV',
    'IfcAirTerminalType': 'RIV',
    'IfcFan': 'RIV',
    'IfcFanType': 'RIV',
    'IfcCoil': 'RIV',
    'IfcCoilType': 'RIV',
    'IfcBoiler': 'RIV',
    'IfcBoilerType': 'RIV',
    'IfcChiller': 'RIV',
    'IfcChillerType': 'RIV',
}


# Column headers and their widths
COLUMNS = [
    # Editable columns (A-D)
    ('NS3451 Code', 15),
    ('Unit', 8),
    ('Notes', 25),
    ('Status', 12),
    # Read-only columns (E-R)
    ('Type Name', 30),
    ('IfcEntity', 20),
    ('IfcType', 25),
    ('Predefined Type', 15),
    ('Discipline', 12),
    ('IsExternal', 12),
    ('LoadBearing', 12),
    ('FireRating', 15),
    ('AcousticRating', 15),
    ('Layer/Storey', 20),
    ('Material', 25),
    ('System', 20),
    ('Instances', 10),
    ('Type GUID', 25),
]


class TypeExcelExporter:
    """Generates Excel template for batch type mapping."""

    def __init__(self, model_id: str):
        self.model_id = model_id
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Type Mapping"

        # Styles
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF')
        self.editable_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        self.readonly_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        self.border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

    def export(self) -> BytesIO:
        """Generate Excel file and return as BytesIO."""
        # Get types with aggregated data
        types_data = self._get_types_with_aggregates()

        # Sort by IfcEntity for grouping
        types_data.sort(key=lambda x: (x.get('ifc_entity', '') or '', x.get('type_name', '') or ''))

        # Write headers
        self._write_headers()

        # Write data rows
        for row_idx, type_data in enumerate(types_data, start=2):
            self._write_row(row_idx, type_data)

        # Setup formatting
        self._setup_column_widths()
        self._setup_data_validation(len(types_data) + 1)
        self._freeze_panes()

        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def _get_types_with_aggregates(self) -> List[Dict[str, Any]]:
        """Query types with all aggregated properties."""
        types = IFCType.objects.filter(model_id=self.model_id).select_related('mapping')

        result = []
        for ifc_type in types:
            # Get instance count and entity info
            assignments = TypeAssignment.objects.filter(type=ifc_type).select_related('entity')
            instance_count = assignments.count()

            # Get entities for aggregation
            entity_ids = list(assignments.values_list('entity_id', flat=True))

            # Aggregate IfcEntity class (most common)
            ifc_entity = self._aggregate_entity_class(entity_ids)

            # Aggregate properties from Psets
            is_external = self._aggregate_property(entity_ids, 'IsExternal')
            load_bearing = self._aggregate_property(entity_ids, 'LoadBearing')
            fire_rating = self._aggregate_property(entity_ids, 'FireRating')
            acoustic_rating = self._aggregate_property(entity_ids, 'AcousticRating')

            # Aggregate storey
            storey = self._aggregate_storey(entity_ids)

            # Aggregate materials
            materials = self._aggregate_materials(entity_ids)

            # Aggregate systems
            systems = self._aggregate_systems(entity_ids)

            # Get existing mapping if any
            mapping = getattr(ifc_type, 'mapping', None)

            # Infer discipline
            discipline = DISCIPLINE_MAP.get(ifc_type.ifc_type, '')
            if mapping and mapping.discipline:
                discipline = mapping.discipline

            result.append({
                'type_guid': ifc_type.type_guid,
                'type_name': ifc_type.type_name or '',
                'ifc_type': ifc_type.ifc_type,
                'ifc_entity': ifc_entity,
                'predefined_type': ifc_type.properties.get('predefined_type', ''),
                'discipline': discipline,
                'is_external': is_external,
                'load_bearing': load_bearing,
                'fire_rating': fire_rating,
                'acoustic_rating': acoustic_rating,
                'storey': storey,
                'materials': materials,
                'systems': systems,
                'instance_count': instance_count,
                # Existing mapping data
                'ns3451_code': mapping.ns3451_code if mapping else '',
                'unit': mapping.representative_unit if mapping else '',
                'notes': mapping.notes if mapping else '',
                'status': mapping.mapping_status if mapping else 'pending',
            })

        return result

    def _aggregate_entity_class(self, entity_ids: List) -> str:
        """Get most common entity class for type instances."""
        if not entity_ids:
            return ''

        entity_types = (
            IFCEntity.objects
            .filter(id__in=entity_ids)
            .values_list('ifc_type', flat=True)
            .distinct()
        )
        types = list(entity_types)

        if len(types) == 1:
            return types[0]
        elif len(types) > 1:
            # Return most common
            counts = {}
            for t in IFCEntity.objects.filter(id__in=entity_ids).values_list('ifc_type', flat=True):
                counts[t] = counts.get(t, 0) + 1
            return max(counts, key=counts.get) if counts else ''
        return ''

    def _aggregate_property(self, entity_ids: List, property_name: str) -> str:
        """Aggregate property value from instances: Yes/No/Mixed or actual value."""
        if not entity_ids:
            return ''

        values = list(
            PropertySet.objects
            .filter(
                entity_id__in=entity_ids,
                property_name=property_name,
            )
            .values_list('property_value', flat=True)
            .distinct()
        )

        if not values:
            return ''

        # For boolean-like properties
        if property_name in ('IsExternal', 'LoadBearing'):
            true_values = {'True', 'true', '1', 'Yes', 'yes'}
            false_values = {'False', 'false', '0', 'No', 'no'}

            has_true = any(v in true_values for v in values if v)
            has_false = any(v in false_values for v in values if v)

            if has_true and has_false:
                return 'Mixed'
            elif has_true:
                return 'Yes'
            elif has_false:
                return 'No'
            return ''

        # For string properties, return most common or list
        if len(values) == 1:
            return values[0] or ''
        elif len(values) <= 3:
            return ', '.join(v for v in values if v)
        else:
            return f"{values[0]} (+{len(values)-1})"

    def _aggregate_storey(self, entity_ids: List) -> str:
        """Aggregate storey/layer info from instances."""
        if not entity_ids:
            return ''

        # Get unique storey names by joining to spatial structure
        from apps.models.models import SpatialElement

        storey_ids = list(
            IFCEntity.objects
            .filter(id__in=entity_ids, storey_id__isnull=False)
            .values_list('storey_id', flat=True)
            .distinct()
        )

        if not storey_ids:
            return ''

        storeys = list(
            SpatialElement.objects
            .filter(id__in=storey_ids)
            .values_list('name', flat=True)
            .distinct()
        )

        if len(storeys) == 1:
            return storeys[0] or ''
        elif len(storeys) <= 3:
            return ', '.join(s for s in storeys if s)
        else:
            return f"{storeys[0]} (+{len(storeys)-1})"

    def _aggregate_materials(self, entity_ids: List) -> str:
        """Aggregate material names from instances."""
        if not entity_ids:
            return ''

        materials = list(
            MaterialAssignment.objects
            .filter(entity_id__in=entity_ids)
            .select_related('material')
            .values_list('material__name', flat=True)
            .distinct()[:5]
        )

        if len(materials) == 1:
            return materials[0] or ''
        elif len(materials) > 1:
            return ', '.join(m for m in materials if m)
        return ''

    def _aggregate_systems(self, entity_ids: List) -> str:
        """Aggregate system names from instances."""
        if not entity_ids:
            return ''

        systems = list(
            SystemMembership.objects
            .filter(entity_id__in=entity_ids)
            .select_related('system')
            .values_list('system__system_name', flat=True)
            .distinct()[:3]
        )

        if len(systems) == 1:
            return systems[0] or ''
        elif len(systems) > 1:
            return ', '.join(s for s in systems if s)
        return ''

    def _write_headers(self):
        """Write header row with styling."""
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

    def _write_row(self, row_idx: int, data: Dict[str, Any]):
        """Write a single data row."""
        row_values = [
            # Editable columns (A-D)
            data.get('ns3451_code', ''),
            data.get('unit', ''),
            data.get('notes', ''),
            data.get('status', 'pending'),
            # Read-only columns (E-R)
            data.get('type_name', ''),
            data.get('ifc_entity', ''),
            data.get('ifc_type', ''),
            data.get('predefined_type', ''),
            data.get('discipline', ''),
            data.get('is_external', ''),
            data.get('load_bearing', ''),
            data.get('fire_rating', ''),
            data.get('acoustic_rating', ''),
            data.get('storey', ''),
            data.get('materials', ''),
            data.get('systems', ''),
            data.get('instance_count', 0),
            data.get('type_guid', ''),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = self.worksheet.cell(row=row_idx, column=col_idx, value=value)

            # Style based on editable vs read-only
            if col_idx <= 4:  # Editable columns A-D
                cell.fill = self.editable_fill
            else:  # Read-only columns E-R
                cell.fill = self.readonly_fill

            cell.border = self.border
            cell.alignment = Alignment(vertical='center')

    def _setup_column_widths(self):
        """Set column widths."""
        for col_idx, (_, width) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            self.worksheet.column_dimensions[col_letter].width = width

    def _setup_data_validation(self, max_row: int):
        """Setup dropdown validation for editable columns."""
        # Unit dropdown (column B)
        unit_validation = DataValidation(
            type='list',
            formula1='"pcs,m,m2,m3"',
            allow_blank=True,
        )
        unit_validation.add(f'B2:B{max_row}')
        self.worksheet.add_data_validation(unit_validation)

        # Status dropdown (column D)
        status_validation = DataValidation(
            type='list',
            formula1='"pending,mapped,ignored,review,followup"',
            allow_blank=True,
        )
        status_validation.add(f'D2:D{max_row}')
        self.worksheet.add_data_validation(status_validation)

        # NS3451 Code validation - just a text input for now
        # (full dropdown would have 200+ options, not practical)

    def _freeze_panes(self):
        """Freeze header row."""
        self.worksheet.freeze_panes = 'A2'


def export_types_to_excel(model_id: str) -> BytesIO:
    """
    Export types for a model to Excel template.

    Args:
        model_id: UUID of the model

    Returns:
        BytesIO containing the Excel file
    """
    exporter = TypeExcelExporter(model_id)
    return exporter.export()
