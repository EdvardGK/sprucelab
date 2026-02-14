"""Excel formatting utilities for IFC analysis reports.

Provides consistent styling (colors, fonts, borders) and helper functions
for openpyxl-based Excel generation. Requires the 'excel' optional dependency.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# -- Color fills --
COLORS = {
    "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "gray": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "light_blue": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "header": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
    "ref": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}

# -- Fonts --
FONTS = {
    "header": Font(bold=True, color="FFFFFF"),
    "bold": Font(bold=True),
    "italic": Font(italic=True),
    "italic_gray": Font(italic=True, color="666666"),
}

# -- Standard border --
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_header_row(ws, headers: list[str], row: int = 1) -> None:
    """Write a styled header row to a worksheet.

    Args:
        ws: openpyxl Worksheet.
        headers: List of column header strings.
        row: Row number (1-indexed, default 1).
    """
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = COLORS["header"]
        cell.font = FONTS["header"]
        cell.border = BORDER


def auto_fit_columns(
    ws,
    min_width: int = 8,
    max_width: int = 50,
    header_row: int = 1,
) -> None:
    """Auto-fit column widths based on cell content.

    Args:
        ws: openpyxl Worksheet.
        min_width: Minimum column width.
        max_width: Maximum column width.
        header_row: Row containing headers (used for initial width estimate).
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, min_width), max_width
        )


def apply_status_fill(cell, status: str) -> None:
    """Apply color fill based on a status string.

    Args:
        cell: openpyxl Cell.
        status: One of 'ok', 'warning', 'critical', 'info', 'neutral'.
    """
    mapping = {
        "ok": COLORS["green"],
        "warning": COLORS["yellow"],
        "critical": COLORS["red"],
        "info": COLORS["light_blue"],
        "neutral": COLORS["gray"],
    }
    fill = mapping.get(status.lower())
    if fill:
        cell.fill = fill


def freeze_and_filter(ws, row: int = 2) -> None:
    """Freeze panes at the given row and enable auto-filter.

    Args:
        ws: openpyxl Worksheet.
        row: Row to freeze at (default 2, freezing row 1 as header).
    """
    ws.freeze_panes = f"A{row}"
    ws.auto_filter.ref = ws.dimensions
